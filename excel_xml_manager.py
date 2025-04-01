import os
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from lxml import etree
import re
import traceback
from openpyxl.utils import get_column_letter
import datetime
import uuid

class ExcelXmlManager:
    """
    Classe per gestire l'interscambio di dati tra file XML delle fatture elettroniche
    e file Excel, permettendo di salvare molteplici fatture con più righe di dettaglio.
    """
    
    def __init__(self, parent, ns):
        """
        Inizializza il manager XML-Excel
        
        Args:
            parent: Riferimento alla classe principale per i log
            ns: Namespace XML da utilizzare
        """
        self.parent = parent
        self.NS = ns
        
        # Aggiungi opzione per specificare il file Excel più tardi
        self.excel_path = None
        
        # Nomi dei fogli Excel
        self.master_sheet_name = "Fatture"
        self.details_sheet_name = "DettaglioLinee"
        self.summary_sheet_name = "DatiRiepilogo"
        self.structure_sheet_name = "StrutturaXML"
    
    def log(self, message):
        """
        Utilizza il metodo di log del parent se disponibile
        """
        if hasattr(self.parent, 'log') and callable(self.parent.log):
            self.parent.log(message)
        else:
            print(message)
    
    def export_xml_to_excel(self, xml_doc, excel_path=None):
        """
        Esporta tutti i dati dal file XML al foglio Excel
        
        Args:
            xml_doc: Documento XML da esportare
            excel_path: Percorso del file Excel (opzionale)
        
        Returns:
            bool: True se l'operazione ha successo, False altrimenti
        """
        if excel_path:
            self.excel_path = excel_path
        # Verifica che il percorso del file Excel sia stato impostato
        if not self.excel_path:
            self.log("Errore: Nessun file Excel specificato. Usa prima 'Carica DB Excel' o 'Crea DB Excel'.")
            return False
            
        try:
            root = xml_doc.getroot()
            
            # Verifica se il file Excel esiste
            if os.path.exists(self.excel_path):
                # Apri il workbook esistente
                wb = openpyxl.load_workbook(self.excel_path)
                self.log(f"File Excel esistente aperto: {self.excel_path}")
            else:
                # Crea un nuovo workbook e rimuovi il foglio di default
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                self.log(f"Nuovo file Excel creato")
            
            # Crea o recupera i fogli necessari
            master_sheet = self._ensure_sheet(wb, self.master_sheet_name)
            details_sheet = self._ensure_sheet(wb, self.details_sheet_name)
            summary_sheet = self._ensure_sheet(wb, self.summary_sheet_name)
            structure_sheet = self._ensure_sheet(wb, self.structure_sheet_name)
            
            # Configura le intestazioni nei fogli se sono vuoti
            self._setup_sheet_headers(master_sheet, [
                "ID_Fattura", "NumeroFattura", "DataFattura", "TipoDocumento", 
                "ImportoTotale", "CedenteDenominazione", "CedentePartitaIVA", 
                "CessionarioDenominazione", "CessionarioPartitaIVA", "NotaFattura"
            ])
            
            self._setup_sheet_headers(details_sheet, [
                "ID_Fattura", "NumeroLinea", "Descrizione", "Quantita", 
                "UnitaMisura", "PrezzoUnitario", "PrezzoTotale", "AliquotaIVA", "Note"
            ])
            
            self._setup_sheet_headers(summary_sheet, [
                "ID_Fattura", "AliquotaIVA", "ImponibileImporto", "Imposta", 
                "EsigibilitaIVA", "Natura"
            ])
            
            self._setup_sheet_headers(structure_sheet, [
                "TagXML", "Percorso", "Descrizione"
            ])
            
            # Genera un ID univoco per questa fattura
            invoice_id = str(uuid.uuid4())
            
            # Estrai i dati principali della fattura
            self.log("Estrazione dati principali della fattura")
            invoice_data = self._extract_invoice_data(root, invoice_id)
            
            # Aggiungi i dati della fattura al foglio principale
            row = master_sheet.max_row + 1
            for col, value in enumerate(invoice_data, 1):
                master_sheet.cell(row=row, column=col, value=value)
            
            # Estrai e salva le linee di dettaglio
            self.log("Estrazione linee di dettaglio")
            detail_lines = self._extract_detail_lines(root, invoice_id)
            
            for line in detail_lines:
                row = details_sheet.max_row + 1
                for col, value in enumerate(line, 1):
                    details_sheet.cell(row=row, column=col, value=value)
            
            # Estrai e salva i dati di riepilogo
            self.log("Estrazione dati di riepilogo")
            summary_data = self._extract_summary_data(root, invoice_id)
            
            for item in summary_data:
                row = summary_sheet.max_row + 1
                for col, value in enumerate(item, 1):
                    summary_sheet.cell(row=row, column=col, value=value)
            
            # Estrai e salva la struttura XML (solo se il foglio è vuoto)
            if structure_sheet.max_row <= 1:
                self.log("Estrazione struttura XML")
                structure_data = self._extract_xml_structure(root)
                
                for item in structure_data:
                    row = structure_sheet.max_row + 1
                    for col, value in enumerate(item, 1):
                        structure_sheet.cell(row=row, column=col, value=value)
            
            # Ottimizza larghezza colonne
            for sheet in [master_sheet, details_sheet, summary_sheet, structure_sheet]:
                self._optimize_column_width(sheet)
            
            # Salva il file Excel
            wb.save(self.excel_path)
            
            self.log(f"Fattura esportata in Excel con ID: {invoice_id}")
            self.log(f"Righe di dettaglio: {len(detail_lines)}")
            self.log(f"Dati di riepilogo: {len(summary_data)}")
            
            return True
        
        except Exception as e:
            self.log(f"Errore nell'esportazione XML in Excel: {str(e)}")
            traceback.print_exc()
            return False
    
    def _ensure_sheet(self, workbook, sheet_name):
        """
        Assicura che il foglio esista, creandolo se necessario
        
        Args:
            workbook: Workbook Excel
            sheet_name: Nome del foglio
        
        Returns:
            Worksheet: Foglio Excel
        """
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        else:
            return workbook.create_sheet(title=sheet_name)
    
    def _setup_sheet_headers(self, sheet, headers):
        """
        Configura le intestazioni in un foglio se è vuoto
        
        Args:
            sheet: Foglio Excel
            headers: Lista di intestazioni
        """
        if sheet.max_row <= 1 and sheet.max_column <= 1:
            # Foglio vuoto, aggiungi intestazioni
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
    
    def _extract_invoice_data(self, root, invoice_id):
        """
        Estrae i dati principali della fattura
        
        Args:
            root: Elemento radice XML
            invoice_id: ID univoco della fattura
        
        Returns:
            list: Dati della fattura
        """
        # Inizializza con valori di default
        numero = ""
        data = ""
        tipo_documento = ""
        importo_totale = ""
        cedente_denominazione = ""
        cedente_partita_iva = ""
        cessionario_denominazione = ""
        cessionario_partita_iva = ""
        
        try:
            # Numero fattura
            elem = root.xpath("//*/DatiGenerali/DatiGeneraliDocumento/Numero", namespaces=self.NS)
            if elem:
                numero = elem[0].text
            
            # Data fattura
            elem = root.xpath("//*/DatiGenerali/DatiGeneraliDocumento/Data", namespaces=self.NS)
            if elem:
                data = elem[0].text
            
            # Tipo documento
            elem = root.xpath("//*/DatiGenerali/DatiGeneraliDocumento/TipoDocumento", namespaces=self.NS)
            if elem:
                tipo_documento = elem[0].text
            
            # Importo totale
            elem = root.xpath("//*/DatiGenerali/DatiGeneraliDocumento/ImportoTotaleDocumento", namespaces=self.NS)
            if elem:
                importo_totale = elem[0].text
            
            # Cedente denominazione
            elem = root.xpath("//*/CedentePrestatore/DatiAnagrafici/Anagrafica/Denominazione", namespaces=self.NS)
            if elem:
                cedente_denominazione = elem[0].text
            
            # Cedente partita IVA
            elem = root.xpath("//*/CedentePrestatore/DatiAnagrafici/IdFiscaleIVA/IdCodice", namespaces=self.NS)
            if elem:
                cedente_partita_iva = elem[0].text
            
            # Cessionario denominazione
            elem = root.xpath("//*/CessionarioCommittente/DatiAnagrafici/Anagrafica/Denominazione", namespaces=self.NS)
            if elem:
                cessionario_denominazione = elem[0].text
            
            # Cessionario partita IVA
            elem = root.xpath("//*/CessionarioCommittente/DatiAnagrafici/IdFiscaleIVA/IdCodice", namespaces=self.NS)
            if elem:
                cessionario_partita_iva = elem[0].text
            
        except Exception as e:
            self.log(f"Errore nell'estrazione dei dati della fattura: {str(e)}")
        
        # Restituisci i dati estratti + ID + colonna note vuota
        return [
            invoice_id, numero, data, tipo_documento,
            importo_totale, cedente_denominazione, cedente_partita_iva,
            cessionario_denominazione, cessionario_partita_iva, ""
        ]
    
    def _extract_detail_lines(self, root, invoice_id):
        """
        Estrae tutte le linee di dettaglio
        
        Args:
            root: Elemento radice XML
            invoice_id: ID univoco della fattura
        
        Returns:
            list: Liste di dati per ciascuna linea
        """
        lines = []
        
        try:
            # Trova tutte le linee di dettaglio
            detail_lines = root.xpath("//*/DatiBeniServizi/DettaglioLinee", namespaces=self.NS)
            
            for line in detail_lines:
                line_data = [invoice_id]  # Inizia con l'ID della fattura
                
                # Estrai i dati di ciascuna linea
                for field in ["NumeroLinea", "Descrizione", "Quantita", "UnitaMisura", 
                             "PrezzoUnitario", "PrezzoTotale", "AliquotaIVA"]:
                    elem = line.find(field)
                    value = elem.text if elem is not None else ""
                    line_data.append(value)
                
                # Aggiungi colonna note vuota
                line_data.append("")
                
                lines.append(line_data)
        
        except Exception as e:
            self.log(f"Errore nell'estrazione delle linee di dettaglio: {str(e)}")
        
        return lines
    
    def _extract_summary_data(self, root, invoice_id):
        """
        Estrae i dati di riepilogo
        
        Args:
            root: Elemento radice XML
            invoice_id: ID univoco della fattura
        
        Returns:
            list: Liste di dati per ciascun riepilogo
        """
        summaries = []
        
        try:
            # Trova tutti i dati di riepilogo
            summary_elements = root.xpath("//*/DatiBeniServizi/DatiRiepilogo", namespaces=self.NS)
            
            for element in summary_elements:
                summary_data = [invoice_id]  # Inizia con l'ID della fattura
                
                # Estrai i dati di ciascun riepilogo
                for field in ["AliquotaIVA", "ImponibileImporto", "Imposta", 
                             "EsigibilitaIVA", "Natura"]:
                    elem = element.find(field)
                    value = elem.text if elem is not None else ""
                    summary_data.append(value)
                
                summaries.append(summary_data)
        
        except Exception as e:
            self.log(f"Errore nell'estrazione dei dati di riepilogo: {str(e)}")
        
        return summaries
    
    def _extract_xml_structure(self, root):
        """
        Estrae la struttura dell'XML per documentazione,
        evitando duplicazioni per elementi ripetuti come DettaglioLinee
        
        Args:
            root: Elemento radice XML
        
        Returns:
            list: Lista di tuple (tag, percorso, descrizione)
        """
        structure = []
        # Teniamo traccia dei percorsi già aggiunti
        added_paths = set()
        
        # Dizionario di descrizioni dei tag
        tag_descriptions = self._get_tag_descriptions()
        
        # Lista di tag che possono apparire più volte con lo stesso parent
        repeatable_tags = ["DettaglioLinee", "DatiRiepilogo"]
        
        # Funzione ricorsiva per esplorare l'albero XML
        def explore_node(node, path="", avoid_children=False):
            tag = self._get_tag_name(node)
            current_path = f"{path}/{tag}" if path else tag
            
            # Se il percorso è già stato aggiunto o deve essere saltato, esci subito
            if current_path in added_paths:
                return
            
            # Aggiungi il nodo corrente alla struttura
            description = tag_descriptions.get(tag, "")
            
            # Aggiungi un'indicazione se il tag è ripetibile
            if tag in repeatable_tags:
                description += " (Elemento ripetibile, gestito in foglio separato)"
            
            structure.append([tag, current_path, description])
            added_paths.add(current_path)
            
            # Non continuare con i figli di un nodo ripetibile per evitare duplicazioni
            if avoid_children:
                return
                
            # Esplora i figli solo se non è un nodo ripetibile
            for child in node:
                child_tag = self._get_tag_name(child)
                
                # Verifica se è un elemento ripetibile
                is_repeatable = child_tag in repeatable_tags
                
                # Se troviamo un elemento ripetibile, lo gestiamo diversamente
                if is_repeatable:
                    # Costruisci il percorso per l'elemento ripetibile
                    child_path = f"{current_path}/{child_tag}"
                    
                    # Lo aggiungiamo solo se non l'abbiamo già fatto
                    if child_path not in added_paths:
                        description = tag_descriptions.get(child_tag, "")
                        description += " (Elemento ripetibile, gestito in foglio separato)"
                        structure.append([child_tag, child_path, description])
                        added_paths.add(child_path)
                        
                        # Esplora il primo figlio per capire la struttura interna
                        # ma evita di aggiungere altri nodi con lo stesso tag
                        explore_node(child, current_path, True)
                else:
                    # Per elementi non ripetibili, esplorazione normale
                    explore_node(child, current_path, False)
        
        try:
            explore_node(root)
        except Exception as e:
            self.log(f"Errore nell'estrazione della struttura XML: {str(e)}")
            traceback.print_exc()
        
        return structure


    def _get_tag_name(self, node):
        """
        Estrae il nome del tag senza namespace
        
        Args:
            node: Nodo XML
        
        Returns:
            str: Nome del tag
        """
        tag = node.tag
        if "}" in tag:
            return tag.split("}")[1]
        return tag
    
    def _get_tag_descriptions(self):
        """
        Restituisce un dizionario di descrizioni per i tag XML
        
        Returns:
            dict: Tag e relative descrizioni
        """
        return {
            "FatturaElettronica": "Documento principale della fattura elettronica",
            "FatturaElettronicaHeader": "Intestazione della fattura",
            "DatiTrasmissione": "Dati trasmissione verso SDI",
            "IdTrasmittente": "Identificativo del trasmittente",
            "IdPaese": "Codice ISO del paese",
            "IdCodice": "Codice identificativo fiscale",
            "ProgressivoInvio": "Numero progressivo di invio",
            "FormatoTrasmissione": "Formato trasmissione (FPR12, FPA12)",
            "CodiceDestinatario": "Codice destinatario SDI",
            "CedentePrestatore": "Fornitore",
            "CessionarioCommittente": "Cliente",
            "DatiAnagrafici": "Dati anagrafici",
            "IdFiscaleIVA": "Identificativo fiscale IVA",
            "CodiceFiscale": "Codice fiscale",
            "Anagrafica": "Dati anagrafici",
            "Denominazione": "Denominazione o ragione sociale",
            "Nome": "Nome persona fisica",
            "Cognome": "Cognome persona fisica",
            "RegimeFiscale": "Regime fiscale",
            "Sede": "Sede legale o amministrativa",
            "Indirizzo": "Indirizzo",
            "NumeroCivico": "Numero civico",
            "CAP": "Codice Avviamento Postale",
            "Comune": "Comune",
            "Provincia": "Provincia (sigla)",
            "Nazione": "Nazione (codice ISO)",
            "FatturaElettronicaBody": "Corpo della fattura",
            "DatiGenerali": "Dati generali del documento",
            "DatiGeneraliDocumento": "Dati generali documento",
            "TipoDocumento": "Tipo documento (TD01, TD02, ...)",
            "Divisa": "Valuta del documento",
            "Data": "Data del documento",
            "Numero": "Numero del documento",
            "ImportoTotaleDocumento": "Importo totale del documento",
            "DatiBeniServizi": "Dati relativi ai beni/servizi",
            "DettaglioLinee": "Dettaglio linee del documento",
            "NumeroLinea": "Numero progressivo della linea",
            "Descrizione": "Descrizione della linea",
            "Quantita": "Quantità",
            "UnitaMisura": "Unità di misura",
            "PrezzoUnitario": "Prezzo unitario",
            "PrezzoTotale": "Prezzo totale",
            "AliquotaIVA": "Aliquota IVA",
            "DatiRiepilogo": "Dati di riepilogo",
            "ImponibileImporto": "Imponibile",
            "Imposta": "Imposta",
            "EsigibilitaIVA": "Esigibilità IVA",
            "DatiPagamento": "Dati del pagamento",
            "CondizioniPagamento": "Condizioni di pagamento",
            "DettaglioPagamento": "Dettaglio del pagamento",
            "ModalitaPagamento": "Modalità di pagamento",
            "DataScadenzaPagamento": "Data scadenza pagamento",
            "ImportoPagamento": "Importo del pagamento"
        }
    
    def _optimize_column_width(self, sheet):
        """
        Ottimizza la larghezza delle colonne in base al contenuto
        
        Args:
            sheet: Foglio Excel
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Imposta larghezza con un po' di padding
            adjusted_width = max_length + 2 if max_length < 50 else 50
            sheet.column_dimensions[column_letter].width = adjusted_width
    

    def _show_invoice_selector(self, master_sheet):
        """
        Mostra un selettore per scegliere la fattura da esportare
        
        Args:
            master_sheet: Foglio principale con le fatture
        
        Returns:
            str: ID della fattura selezionata o None
        """
        try:
            # Log per debugging
            self.log("Avvio selettore fatture")
            
            # Estrai l'elenco delle fatture
            invoices = []
            for row in range(2, master_sheet.max_row + 1):
                invoice_id = master_sheet.cell(row=row, column=1).value
                if not invoice_id:
                    continue
                    
                numero = master_sheet.cell(row=row, column=2).value or ""
                data = master_sheet.cell(row=row, column=3).value or ""
                cedente = master_sheet.cell(row=row, column=6).value or ""
                cessionario = master_sheet.cell(row=row, column=8).value or ""
                
                # Log dei valori per debug
                self.log(f"Fattura trovata: ID={invoice_id}, Numero={numero}, Data={data}")
                
                display_text = f"{numero} - {data} - {cedente} → {cessionario}"
                invoices.append((invoice_id, display_text))
            
            if not invoices:
                self.log("Nessuna fattura trovata nel foglio Excel")
                messagebox.showinfo("Informazione", "Nessuna fattura trovata nel foglio Excel")
                return None
            
            # Log per debug
            self.log(f"Trovate {len(invoices)} fatture")
            
            # Variabile per memorizzare l'ID selezionato
            selected_id = None
            
            # Crea una finestra di dialogo per la selezione
            selector = tk.Toplevel(self.parent)
            selector.title("Seleziona Fattura")
            selector.geometry("600x400")
            selector.transient(self.parent)
            selector.grab_set()
            selector.focus_set()
            
            tk.Label(selector, text="Seleziona la fattura da esportare:", 
                    pady=10, font=("", 10, "bold")).pack(fill=tk.X)
            
            # Frame con listbox e scrollbar
            list_frame = tk.Frame(selector)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            scrollbar = tk.Scrollbar(list_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            listbox = tk.Listbox(list_frame, font=("", 10), height=15, 
                            yscrollcommand=scrollbar.set)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            scrollbar.config(command=listbox.yview)
            
            # Importante: memorizza gli ID in una lista separata per un accesso sicuro
            invoice_ids = []
            
            # Popola la listbox
            for i, (invoice_id, display_text) in enumerate(invoices):
                listbox.insert(tk.END, display_text)
                invoice_ids.append(invoice_id)  # Memorizza l'ID nella lista
            
            # Frame per i pulsanti
            btn_frame = tk.Frame(selector)
            btn_frame.pack(fill=tk.X, padx=10, pady=10)
            
            def on_select():
                nonlocal selected_id
                selection = listbox.curselection()
                if selection:
                    idx = selection[0]
                    if 0 <= idx < len(invoice_ids):
                        selected_id = invoice_ids[idx]  # Usa l'ID dalla lista
                        self.log(f"Fattura selezionata: {selected_id}")
                        selector.destroy()
                
            def on_cancel():
                nonlocal selected_id
                selected_id = None
                self.log("Selezione fattura annullata")
                selector.destroy()
            
            # Pulsanti
            select_btn = tk.Button(btn_frame, text="Seleziona", command=on_select,
                                bg="#4CAF50", fg="white", width=15, pady=5)
            select_btn.pack(side=tk.RIGHT, padx=5)
            
            cancel_btn = tk.Button(btn_frame, text="Annulla", command=on_cancel,
                                width=15, pady=5)
            cancel_btn.pack(side=tk.RIGHT, padx=5)
            
            # Doppio click per selezionare
            listbox.bind("<Double-1>", lambda e: on_select())
            
            # Attendi che la finestra venga chiusa
            selector.wait_window()
            
            # Log dell'ID selezionato
            self.log(f"ID selezionato finale: {selected_id}")
            
            return selected_id
            
        except Exception as e:
            self.log(f"Errore nel selettore fatture: {str(e)}")
            traceback.print_exc()
            messagebox.showerror("Errore", f"Errore nella selezione della fattura:\n{str(e)}")
            return None

    def _get_invoice_data_by_id(self, workbook, invoice_id):
        """
        Estrae tutti i dati di una fattura specifica
        
        Args:
            workbook: Workbook Excel
            invoice_id: ID della fattura
        
        Returns:
            dict: Dati della fattura
        """
        invoice_data = {
            "master": None,
            "details": [],
            "summary": [],
            "structure": self._extract_xml_structure_from_sheet(workbook[self.structure_sheet_name])
        }
        
        # Estrai i dati principali della fattura
        master_sheet = workbook[self.master_sheet_name]
        for row in range(2, master_sheet.max_row + 1):
            if master_sheet.cell(row=row, column=1).value == invoice_id:
                invoice_data["master"] = [
                    master_sheet.cell(row=row, column=i).value 
                    for i in range(1, master_sheet.max_column + 1)
                ]
                break
        
        if not invoice_data["master"]:
            return None
        
 
        # Estrai le linee di dettaglio
        details_sheet = workbook[self.details_sheet_name]
        for row in range(2, details_sheet.max_row + 1):
            if details_sheet.cell(row=row, column=1).value == invoice_id:
                line_data = [details_sheet.cell(row=row, column=i).value 
                            for i in range(1, details_sheet.max_column + 1)]
                invoice_data["details"].append(line_data)
        
        # Estrai i dati di riepilogo
        summary_sheet = workbook[self.summary_sheet_name]
        for row in range(2, summary_sheet.max_row + 1):
            if summary_sheet.cell(row=row, column=1).value == invoice_id:
                summary_data = [summary_sheet.cell(row=row, column=i).value 
                                for i in range(1, summary_sheet.max_column + 1)]
                invoice_data["summary"].append(summary_data)
        
        return invoice_data


    def _extract_xml_structure_from_sheet(self, structure_sheet):
        """
        Estrae la struttura XML dal foglio dedicato
        
        Args:
            structure_sheet: Foglio Excel con la struttura
        
        Returns:
            list: Lista di tuple (tag, percorso, descrizione)
        """
        structure = []
        
        for row in range(2, structure_sheet.max_row + 1):
            tag = structure_sheet.cell(row=row, column=1).value
            path = structure_sheet.cell(row=row, column=2).value
            description = structure_sheet.cell(row=row, column=3).value
            
            if tag and path:
                # Aggiungi solo se non è un duplicato di elementi ripetibili
                # (evita percorsi con [n] o simili)
                if not any(item[1] == path for item in structure):
                    structure.append((tag, path, description or ""))
        
        return structure

    def _generate_xml_from_invoice_data(self, invoice_data):
        """
        Genera un documento XML dai dati della fattura
        
        Args:
            invoice_data: Dati della fattura estratti da Excel
        
        Returns:
            etree.ElementTree: Documento XML generato
        """
        ns_uri = "http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2"
        
        # Crea elemento radice SENZA prefisso nel namespace
        nsmap = {None: ns_uri}
        root = etree.Element(f"{{{ns_uri}}}FatturaElettronica", nsmap=nsmap)
        root.set("versione", "FPR12")
        
        # Crea la struttura base
        header = etree.SubElement(root, f"{{{ns_uri}}}FatturaElettronicaHeader")
        body = etree.SubElement(root, f"{{{ns_uri}}}FatturaElettronicaBody")
        
        # Ottieni i dati principali
        master_data = invoice_data["master"]
        details_data = invoice_data["details"]
        summary_data = invoice_data["summary"]
        
        # Crea struttura DatiTrasmissione
        dati_trasmissione = etree.SubElement(header, f"{{{ns_uri}}}DatiTrasmissione")
        id_trasmittente = etree.SubElement(dati_trasmissione, f"{{{ns_uri}}}IdTrasmittente")
        etree.SubElement(id_trasmittente, f"{{{ns_uri}}}IdPaese").text = "IT"  # Default
        etree.SubElement(id_trasmittente, f"{{{ns_uri}}}IdCodice").text = master_data[6] or "00000000000"  # Usa P.IVA cedente
        etree.SubElement(dati_trasmissione, f"{{{ns_uri}}}ProgressivoInvio").text = master_data[1] or "00001"  # Usa numero fattura
        etree.SubElement(dati_trasmissione, f"{{{ns_uri}}}FormatoTrasmissione").text = "FPR12"  # Default
        etree.SubElement(dati_trasmissione, f"{{{ns_uri}}}CodiceDestinatario").text = "0000000"  # Default
        
        # Crea struttura CedentePrestatore
        cedente = etree.SubElement(header, f"{{{ns_uri}}}CedentePrestatore")
        dati_anagrafici_cedente = etree.SubElement(cedente, f"{{{ns_uri}}}DatiAnagrafici")
        id_fiscale_iva_cedente = etree.SubElement(dati_anagrafici_cedente, f"{{{ns_uri}}}IdFiscaleIVA")
        etree.SubElement(id_fiscale_iva_cedente, f"{{{ns_uri}}}IdPaese").text = "IT"  # Default
        etree.SubElement(id_fiscale_iva_cedente, f"{{{ns_uri}}}IdCodice").text = master_data[6] or "00000000000"  # P.IVA cedente
        anagrafica_cedente = etree.SubElement(dati_anagrafici_cedente, f"{{{ns_uri}}}Anagrafica")
        etree.SubElement(anagrafica_cedente, f"{{{ns_uri}}}Denominazione").text = master_data[5] or "Denominazione Cedente"
        etree.SubElement(dati_anagrafici_cedente, f"{{{ns_uri}}}RegimeFiscale").text = "RF01"  # Default
        
        sede_cedente = etree.SubElement(cedente, f"{{{ns_uri}}}Sede")
        etree.SubElement(sede_cedente, f"{{{ns_uri}}}Indirizzo").text = "Indirizzo"  # Default
        etree.SubElement(sede_cedente, f"{{{ns_uri}}}CAP").text = "00000"  # Default
        etree.SubElement(sede_cedente, f"{{{ns_uri}}}Comune").text = "Comune"  # Default
        etree.SubElement(sede_cedente, f"{{{ns_uri}}}Provincia").text = "RM"  # Default
        etree.SubElement(sede_cedente, f"{{{ns_uri}}}Nazione").text = "IT"  # Default
        
        # Crea struttura CessionarioCommittente
        cessionario = etree.SubElement(header, f"{{{ns_uri}}}CessionarioCommittente")
        dati_anagrafici_cessionario = etree.SubElement(cessionario, f"{{{ns_uri}}}DatiAnagrafici")
        
        # Aggiungi partita IVA se presente
        if master_data[8]:
            id_fiscale_iva_cessionario = etree.SubElement(dati_anagrafici_cessionario, f"{{{ns_uri}}}IdFiscaleIVA")
            etree.SubElement(id_fiscale_iva_cessionario, f"{{{ns_uri}}}IdPaese").text = "IT"  # Default
            etree.SubElement(id_fiscale_iva_cessionario, f"{{{ns_uri}}}IdCodice").text = master_data[8]
        
        # Aggiungi sempre anagrafica
        anagrafica_cessionario = etree.SubElement(dati_anagrafici_cessionario, f"{{{ns_uri}}}Anagrafica")
        etree.SubElement(anagrafica_cessionario, f"{{{ns_uri}}}Denominazione").text = master_data[7] or "Denominazione Cessionario"
        
        sede_cessionario = etree.SubElement(cessionario, f"{{{ns_uri}}}Sede")
        etree.SubElement(sede_cessionario, f"{{{ns_uri}}}Indirizzo").text = "Indirizzo"  # Default
        etree.SubElement(sede_cessionario, f"{{{ns_uri}}}CAP").text = "00000"  # Default
        etree.SubElement(sede_cessionario, f"{{{ns_uri}}}Comune").text = "Comune"  # Default
        etree.SubElement(sede_cessionario, f"{{{ns_uri}}}Provincia").text = "RM"  # Default
        etree.SubElement(sede_cessionario, f"{{{ns_uri}}}Nazione").text = "IT"  # Default
        
        # Crea struttura DatiGenerali
        dati_generali = etree.SubElement(body, f"{{{ns_uri}}}DatiGenerali")
        dati_generali_documento = etree.SubElement(dati_generali, f"{{{ns_uri}}}DatiGeneraliDocumento")
        etree.SubElement(dati_generali_documento, f"{{{ns_uri}}}TipoDocumento").text = master_data[3] or "TD01"
        etree.SubElement(dati_generali_documento, f"{{{ns_uri}}}Divisa").text = "EUR"  # Default
        
        # Data fattura
        data_fattura = master_data[2]
        if isinstance(data_fattura, datetime.datetime):
            data_fattura = data_fattura.strftime("%Y-%m-%d")
        etree.SubElement(dati_generali_documento, f"{{{ns_uri}}}Data").text = data_fattura or datetime.date.today().strftime("%Y-%m-%d")
        
        etree.SubElement(dati_generali_documento, f"{{{ns_uri}}}Numero").text = master_data[1] or "00001"
        etree.SubElement(dati_generali_documento, f"{{{ns_uri}}}ImportoTotaleDocumento").text = str(master_data[4] or "0.00")
        
        # Crea struttura DatiBeniServizi
        dati_beni = etree.SubElement(body, f"{{{ns_uri}}}DatiBeniServizi")
        
        # Aggiungi DettaglioLinee
        for line_data in details_data:
            dettaglio = etree.SubElement(dati_beni, f"{{{ns_uri}}}DettaglioLinee")
            etree.SubElement(dettaglio, f"{{{ns_uri}}}NumeroLinea").text = str(line_data[1] or "1")
            etree.SubElement(dettaglio, f"{{{ns_uri}}}Descrizione").text = str(line_data[2] or "Descrizione")
            
            # Gestisci i campi opzionali
            if line_data[3]:  # Quantità
                etree.SubElement(dettaglio, f"{{{ns_uri}}}Quantita").text = str(line_data[3])
            
            if line_data[4]:  # Unità misura
                etree.SubElement(dettaglio, f"{{{ns_uri}}}UnitaMisura").text = str(line_data[4])
            
            etree.SubElement(dettaglio, f"{{{ns_uri}}}PrezzoUnitario").text = str(line_data[5] or "0.00")
            etree.SubElement(dettaglio, f"{{{ns_uri}}}PrezzoTotale").text = str(line_data[6] or "0.00")
            etree.SubElement(dettaglio, f"{{{ns_uri}}}AliquotaIVA").text = str(line_data[7] or "22.00")
        
        # Aggiungi DatiRiepilogo
        for summary_item in summary_data:
            riepilogo = etree.SubElement(dati_beni, f"{{{ns_uri}}}DatiRiepilogo")
            etree.SubElement(riepilogo, f"{{{ns_uri}}}AliquotaIVA").text = str(summary_item[1] or "22.00")
            etree.SubElement(riepilogo, f"{{{ns_uri}}}ImponibileImporto").text = str(summary_item[2] or "0.00")
            etree.SubElement(riepilogo, f"{{{ns_uri}}}Imposta").text = str(summary_item[3] or "0.00")
            
            # Aggiungi campi opzionali
            if summary_item[4]:  # EsigibilitaIVA
                etree.SubElement(riepilogo, f"{{{ns_uri}}}EsigibilitaIVA").text = str(summary_item[4])
            
            if summary_item[5]:  # Natura
                etree.SubElement(riepilogo, f"{{{ns_uri}}}Natura").text = str(summary_item[5])
        
        # Crea struttura DatiPagamento
        dati_pagamento = etree.SubElement(body, f"{{{ns_uri}}}DatiPagamento")
        etree.SubElement(dati_pagamento, f"{{{ns_uri}}}CondizioniPagamento").text = "TP02"  # Default
        
        dettaglio_pagamento = etree.SubElement(dati_pagamento, f"{{{ns_uri}}}DettaglioPagamento")
        etree.SubElement(dettaglio_pagamento, f"{{{ns_uri}}}ModalitaPagamento").text = "MP05"  # Default
        
        # Data scadenza (30 giorni dalla data fattura)
        data_fattura_obj = None
        try:
            if isinstance(data_fattura, str):
                data_fattura_obj = datetime.datetime.strptime(data_fattura, "%Y-%m-%d").date()
            elif isinstance(data_fattura, datetime.datetime):
                data_fattura_obj = data_fattura.date()
        except:
            data_fattura_obj = datetime.date.today()
        
        if not data_fattura_obj:
            data_fattura_obj = datetime.date.today()
            
        data_scadenza = data_fattura_obj + datetime.timedelta(days=30)
        etree.SubElement(dettaglio_pagamento, f"{{{ns_uri}}}DataScadenzaPagamento").text = data_scadenza.strftime("%Y-%m-%d")
        
        etree.SubElement(dettaglio_pagamento, f"{{{ns_uri}}}ImportoPagamento").text = str(master_data[4] or "0.00")
        
        # Applica indentazione per migliorare la leggibilità
        self._indent_xml(root)
        
        # Crea l'albero XML
        tree = etree.ElementTree(root)
        
        # Aggiungi riferimento allo stylesheet XSL
        pi = etree.ProcessingInstruction('xml-stylesheet', 'type="text/xsl" href="./fatturapa_v1.2_asw.xsl"')
        tree.getroot().addprevious(pi)
        
        return tree

    def _indent_xml(self, elem, level=0):
        """
        Applica indentazione per formattare l'XML in modo leggibile
        
        Args:
            elem: Elemento XML
            level: Livello di indentazione
        """
        i = "\n" + "  " * level
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            for child in elem:
                self._indent_xml(child, level + 1)
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
        else:
            if not elem.tail or not elem.tail.strip():
                elem.tail = i

    def import_excel_to_xml(self, template_xml_path=None, output_xml_path=None):
        """
        Crea un nuovo XML a partire dai dati in Excel (mostra selettore fattura)
        
        Args:
            template_xml_path: Percorso di un file XML modello (opzionale)
            output_xml_path: Percorso di output del nuovo XML (opzionale)
        
        Returns:
            bool, str: (Successo, Percorso del file creato)
        """
        try:
            self.log("Avviata funzione import_excel_to_xml")
            
            # Verifica che il file Excel esista
            if not os.path.exists(self.excel_path):
                self.log(f"File Excel non trovato: {self.excel_path}")
                messagebox.showerror("Errore", f"File Excel non trovato: {self.excel_path}")
                return False, ""
            
            # Carica il workbook
            wb = openpyxl.load_workbook(self.excel_path)
            
            # Verifica che i fogli necessari esistano
            required_sheets = [self.master_sheet_name, self.details_sheet_name, 
                            self.summary_sheet_name, self.structure_sheet_name]
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    self.log(f"Foglio '{sheet_name}' non trovato nel file Excel")
                    messagebox.showerror("Errore", f"Foglio '{sheet_name}' non trovato nel file Excel")
                    return False, ""
            
            # Mostra il selettore per scegliere la fattura
            invoice_id = self._show_invoice_selector(wb[self.master_sheet_name])
            if not invoice_id:
                self.log("Nessuna fattura selezionata")
                return False, ""
            
            # Se non è specificato un percorso di output, chiedi all'utente
            if not output_xml_path:
                output_xml_path = filedialog.asksaveasfilename(
                    title="Salva il nuovo file XML",
                    defaultextension=".xml",
                    filetypes=[("File XML", "*.xml")]
                )
                if not output_xml_path:
                    self.log("Operazione di creazione XML annullata dall'utente")
                    return False, ""
            
            # Estrai i dati della fattura
            invoice_data = self._get_invoice_data_by_id(wb, invoice_id)
            if not invoice_data:
                self.log(f"Dati non trovati per la fattura con ID: {invoice_id}")
                messagebox.showerror("Errore", f"Dati non trovati per la fattura con ID: {invoice_id}")
                return False, ""
            
            # Crea il documento XML
            xml_doc = self._generate_xml_from_invoice_data(invoice_data)
            
            # Salva il file XML
            xml_string = etree.tostring(xml_doc, pretty_print=True, encoding="UTF-8", 
                                        xml_declaration=True).decode("utf-8")
            
            # Migliora formattazione per nodi ripetuti
            xml_string = re.sub(r'(</DettaglioLinee>)(\r?\n)+(<(?:\w+:)?DatiRiepilogo>)', 
                            r'\1\n      \3', xml_string)
            xml_string = re.sub(r'(</DettaglioLinee>)(<(?:\w+:)?DatiRiepilogo>)', 
                            r'\1\n      \2', xml_string)
            
            # Aggiungi il prefisso "p:" SOLO all'elemento radice FatturaElettronica
            xml_string = xml_string.replace("<FatturaElettronica ", "<p:FatturaElettronica ")
            xml_string = xml_string.replace("</FatturaElettronica>", "</p:FatturaElettronica>")
            
            # Aggiungi il namespace "xmlns:p" all'elemento radice
            xml_string = xml_string.replace("<p:FatturaElettronica ", 
                                        "<p:FatturaElettronica xmlns:p=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\" ")
            
            # Rimuovi il namespace senza prefisso, per evitare di avere entrambi
            xml_string = xml_string.replace(" xmlns=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\"", "")
            
            with open(output_xml_path, 'w', encoding='utf-8') as f:
                f.write(xml_string)
            
            self.log(f"File XML creato con successo: {output_xml_path}")
            return True, output_xml_path
            
        except Exception as e:
            self.log(f"Errore nella creazione del file XML: {str(e)}")
            traceback.print_exc()
            messagebox.showerror("Errore", f"Errore nella creazione del file XML:\n{str(e)}")
            return False, ""



    def list_invoices(self):
        """
        Restituisce un elenco delle fatture presenti nel file Excel
        
        Returns:
            list: Lista di tuple (id, numero, data, cedente, cessionario)
        """
        try:
            # Verifica che il file Excel esista
            if not os.path.exists(self.excel_path):
                self.log(f"File Excel non trovato: {self.excel_path}")
                return []
            
            # Carica il workbook in modalità sola lettura
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            
            # Verifica che il foglio master esista
            if self.master_sheet_name not in wb.sheetnames:
                self.log(f"Foglio '{self.master_sheet_name}' non trovato nel file Excel")
                return []
            
            # Estrai l'elenco delle fatture
            invoices = []
            master_sheet = wb[self.master_sheet_name]
            
            # Ottieni gli indici delle colonne (per supportare anche fogli con colonne diverse)
            headers = next(master_sheet.rows)
            col_indices = {cell.value: i for i, cell in enumerate(headers)}
            
            # Verifica che ci siano le colonne necessarie
            required_cols = ["ID_Fattura", "NumeroFattura", "DataFattura", 
                            "CedenteDenominazione", "CessionarioDenominazione"]
            
            for col in required_cols:
                if col not in col_indices:
                    self.log(f"Colonna '{col}' non trovata nel foglio principale")
                    return []
            
            # Estrai i dati
            for row in list(master_sheet.rows)[1:]:  # Salta intestazione
                invoice_id = row[col_indices["ID_Fattura"]].value
                if not invoice_id:
                    continue
                    
                numero = row[col_indices["NumeroFattura"]].value or ""
                data = row[col_indices["DataFattura"]].value or ""
                
                # Formatta la data se necessario
                if isinstance(data, datetime.datetime):
                    data = data.strftime("%Y-%m-%d")
                    
                cedente = row[col_indices["CedenteDenominazione"]].value or ""
                cessionario = row[col_indices["CessionarioDenominazione"]].value or ""
                
                invoices.append((invoice_id, numero, data, cedente, cessionario))
            
            return invoices
        
        except Exception as e:
            self.log(f"Errore nell'elenco delle fatture: {str(e)}")
            traceback.print_exc()
            return []

    def delete_invoice(self, invoice_id):
        """
        Elimina una fattura dal file Excel
        
        Args:
            invoice_id: ID della fattura da eliminare
        
        Returns:
            bool: True se l'operazione ha successo, False altrimenti
        """
        try:
            # Verifica che il file Excel esista
            if not os.path.exists(self.excel_path):
                self.log(f"File Excel non trovato: {self.excel_path}")
                return False
            
            # Carica il workbook
            wb = openpyxl.load_workbook(self.excel_path)
            
            # Verifica che i fogli necessari esistano
            required_sheets = [self.master_sheet_name, self.details_sheet_name, self.summary_sheet_name]
            missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
            
            if missing_sheets:
                self.log(f"Fogli mancanti: {', '.join(missing_sheets)}")
                return False
            
            # Elimina le righe corrispondenti nei vari fogli
            rows_deleted = 0
            
            # Elimina dal foglio principale
            rows_to_delete = []
            master_sheet = wb[self.master_sheet_name]
            
            for row_idx in range(master_sheet.max_row, 1, -1):  # Inizia dal fondo
                if master_sheet.cell(row=row_idx, column=1).value == invoice_id:
                    rows_to_delete.append(row_idx)
            
            for row_idx in rows_to_delete:
                master_sheet.delete_rows(row_idx)
                rows_deleted += 1
            
            self.log(f"Rimosse {len(rows_to_delete)} righe dal foglio principale")
            
            # Elimina dal foglio dettagli
            rows_to_delete = []
            details_sheet = wb[self.details_sheet_name]
            
            for row_idx in range(details_sheet.max_row, 1, -1):  # Inizia dal fondo
                if details_sheet.cell(row=row_idx, column=1).value == invoice_id:
                    rows_to_delete.append(row_idx)
            
            for row_idx in rows_to_delete:
                details_sheet.delete_rows(row_idx)
                rows_deleted += 1
            
            self.log(f"Rimosse {len(rows_to_delete)} righe dal foglio dettagli")
            
            # Elimina dal foglio riepilogo
            rows_to_delete = []
            summary_sheet = wb[self.summary_sheet_name]
            
            for row_idx in range(summary_sheet.max_row, 1, -1):  # Inizia dal fondo
                if summary_sheet.cell(row=row_idx, column=1).value == invoice_id:
                    rows_to_delete.append(row_idx)
            
            for row_idx in rows_to_delete:
                summary_sheet.delete_rows(row_idx)
                rows_deleted += 1
            
            self.log(f"Rimosse {len(rows_to_delete)} righe dal foglio riepilogo")
            
            # Salva il file Excel
            wb.save(self.excel_path)
            
            self.log(f"Fattura con ID {invoice_id} eliminata. Totale righe rimosse: {rows_deleted}")
            return rows_deleted > 0
        
        except Exception as e:
            self.log(f"Errore nell'eliminazione della fattura: {str(e)}")
            traceback.print_exc()
            return False
        


    def create_xml_from_excel_by_id(self, invoice_id=None, output_xml_path=None):
        """
        Crea un nuovo XML a partire dai dati in Excel per un ID fattura specifico
        
        Args:
            invoice_id: ID della fattura da creare (se None, viene mostrato un selettore)
            output_xml_path: Percorso di output del nuovo XML (opzionale)
        
        Returns:
            bool, str: (Successo, Percorso del file creato)
        """
        try:
            # Verifica che il file Excel esista
            if not self.excel_path:
                self.log(f"File Excel non specificato")
                return False, ""
                
            if not os.path.exists(self.excel_path):
                self.log(f"File Excel non trovato: {self.excel_path}")
                return False, ""
            
            # Carica il workbook
            wb = openpyxl.load_workbook(self.excel_path)
            
            # Verifica che i fogli necessari esistano
            required_sheets = [self.master_sheet_name, self.details_sheet_name, 
                            self.summary_sheet_name, self.structure_sheet_name]
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    self.log(f"Foglio '{sheet_name}' non trovato nel file Excel")
                    return False, ""
            
            # Se non è specificato un ID e non è un percorso temporaneo, 
            # mostra il selettore di fatture
            if invoice_id is None and (output_xml_path is None or 
                                    "temp_invoice_" not in output_xml_path):
                invoice_id = self._show_invoice_selector(wb[self.master_sheet_name])
                if not invoice_id:
                    self.log("Nessuna fattura selezionata")
                    return False, ""
            
            # Se non è specificato un percorso di output, chiedi all'utente
            if not output_xml_path:
                output_xml_path = filedialog.asksaveasfilename(
                    title="Salva il nuovo file XML",
                    defaultextension=".xml",
                    filetypes=[("File XML", "*.xml")]
                )
                if not output_xml_path:
                    self.log("Operazione di creazione XML annullata dall'utente")
                    return False, ""
            
            # Estrai i dati della fattura
            invoice_data = self._get_invoice_data_by_id(wb, invoice_id)
            if not invoice_data:
                self.log(f"Dati non trovati per la fattura con ID: {invoice_id}")
                return False, ""
            
            # Crea il documento XML
            xml_doc = self._generate_xml_from_invoice_data(invoice_data)
            
            # Salva il file XML
            xml_string = etree.tostring(xml_doc, pretty_print=True, encoding="UTF-8", 
                                        xml_declaration=True).decode("utf-8")
            
            # Migliora formattazione per nodi ripetuti
            xml_string = re.sub(r'(</DettaglioLinee>)(\r?\n)+(<(?:\w+:)?DatiRiepilogo>)', 
                            r'\1\n      \3', xml_string)
            xml_string = re.sub(r'(</DettaglioLinee>)(<(?:\w+:)?DatiRiepilogo>)', 
                            r'\1\n      \2', xml_string)
            
            # Aggiungi il prefisso "p:" SOLO all'elemento radice FatturaElettronica
            xml_string = xml_string.replace("<FatturaElettronica ", "<p:FatturaElettronica ")
            xml_string = xml_string.replace("</FatturaElettronica>", "</p:FatturaElettronica>")
            
            # Aggiungi il namespace "xmlns:p" all'elemento radice
            xml_string = xml_string.replace("<p:FatturaElettronica ", 
                                        "<p:FatturaElettronica xmlns:p=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\" ")
            
            # Rimuovi il namespace senza prefisso, per evitare di avere entrambi
            xml_string = xml_string.replace(" xmlns=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\"", "")
            
            # Aggiungi il riferimento allo stylesheet XSL
            stylesheet_ref = '<?xml-stylesheet type="text/xsl" href="./fatturapa_v1.2_asw.xsl"?>\n'
            if '<?xml ' in xml_string:
                xml_decl_end = xml_string.find('?>') + 2
                xml_string = xml_string[:xml_decl_end] + '\n' + stylesheet_ref + xml_string[xml_decl_end:]
            else:
                xml_string = stylesheet_ref + xml_string
            
            with open(output_xml_path, 'w', encoding='utf-8') as f:
                f.write(xml_string)
            
            self.log(f"File XML creato con successo: {output_xml_path}")
            return True, output_xml_path
            
        except Exception as e:
            self.log(f"Errore nella creazione del file XML: {str(e)}")
            traceback.print_exc()
            return False, ""