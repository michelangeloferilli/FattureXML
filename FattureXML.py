from excel_xml_manager import ExcelXmlManager
from autocomplete_comuni import AutocompleteComune
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import os
import tempfile
import webbrowser
from lxml import etree
import sys
import glob
import copy
import traceback
import re
from tkcalendar import DateEntry
import datetime

class FatturaViewer(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gestione Fatture Elettroniche")
        self.geometry("960x640")
        
        self.xml_path = None
        self.xsl_path = None
        self.xsl_files = []
        self.project_dir = os.path.dirname(os.path.abspath(__file__))
        self.xml_doc = None  # Documento XML caricato
        self.edit_widgets = []  # Widget dell'editor
        # Dizionario per memorizzare le modifiche sui campi della linea di dettaglio:
        # chiave: indice della linea, valore: dict {xpath: nuovo_valore}
        self.line_modifications = {}
        

        # Riferimenti ai pulsanti per gestire lo stato abilitato/disabilitato
        self.edit_btn = None
        self.view_btn = None
        self.excel_save_btn = None
        
        # Inizializza le variabili per le linee
        self.normal_lines = []
        self.conai_line = None
        self.total_lines = 0
        self.current_line_index = 0        
        # Definizione statica del namespace:
        # Anche se nel file XML gli elementi non mostrano il prefisso, questi sono comunque in questo namespace.
        self.NS = {"p": "http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2"}
        # Inizializza il manager Excel-XML
        self.excel_manager = ExcelXmlManager(self, self.NS)
        self.excel_manage_btn = None  # Pulsante Gestisci Fatture
        self.excel_db_label = None    # Etichetta per mostrare il db corrente        
        
        self.create_widgets()
        self.find_xsl_files()

    def indent(self, elem, level=0):
        """Applica indentazione ricorsiva all'albero XML per una formattazione leggibile."""
        i = "\n" + "  " * level
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            for child in elem:
                self.indent(child, level + 1)
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
        else:
            if not elem.tail or not elem.tail.strip():
                elem.tail = i



    def create_date_field(self, parent, value, width=25):
        """
        Crea un campo data con datepicker
        
        Args:
            parent: Widget genitore in cui inserire il campo
            value: Valore iniziale della data (formato YYYY-MM-DD)
            width: Larghezza del campo di input
            
        Returns:
            tuple: (frame_contenitore, widget_entry)
        """
        # Frame che conterrà sia l'entry che il pulsante calendario
        date_frame = tk.Frame(parent)
        
        # Campo di input per la data
        entry_widget = tk.Entry(date_frame, width=width)
        entry_widget.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Inserisci la data iniziale nel campo
        entry_widget.insert(0, value)
        
        # Funzione per aprire il datepicker
        def open_calendar(event=None):
            # Ottieni la posizione corrente del puntatore
            x, y = self.winfo_pointerxy()
            
            # Crea una finestra popup per il calendario
            top = tk.Toplevel(self)
            top.geometry(f"+{x}+{y}")  # Posiziona vicino al puntatore
            top.title("Seleziona data")
            top.resizable(False, False)
            top.grab_set()  # Rende la finestra modale
            
            # Tenta di usare la data già presente nel campo come default
            try:
                if entry_widget.get():
                    date_obj = datetime.datetime.strptime(entry_widget.get(), "%Y-%m-%d").date()
                else:
                    date_obj = datetime.date.today()
            except ValueError:
                date_obj = datetime.date.today()
            
            # Crea il widget calendario
            cal = DateEntry(top, width=12, background='darkblue',
                            foreground='white', borderwidth=2, 
                            date_pattern='yyyy-mm-dd',
                            year=date_obj.year, 
                            month=date_obj.month, 
                            day=date_obj.day)
            cal.pack(padx=10, pady=10)
            
            # Pulsanti di conferma e annullamento
            btn_frame = tk.Frame(top)
            btn_frame.pack(fill="x", padx=10, pady=5)
            
            # Funzione per impostare la data selezionata nel campo di input
            def set_date():
                selected_date = cal.get_date().strftime("%Y-%m-%d")
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, selected_date)
                top.destroy()
            
            # Pulsante OK
            ok_btn = tk.Button(btn_frame, text="OK", command=set_date)
            ok_btn.pack(side=tk.RIGHT, padx=5)
            
            # Pulsante Annulla
            cancel_btn = tk.Button(btn_frame, text="Annulla", command=top.destroy)
            cancel_btn.pack(side=tk.RIGHT, padx=5)
        
        # Collega l'azione di doppio clic all'apertura del calendario
        entry_widget.bind("<Double-1>", open_calendar)
        
        # Aggiungi un pulsante con icona calendario
        calendar_btn = tk.Button(date_frame, text="📅", command=open_calendar)
        calendar_btn.pack(side=tk.RIGHT)
        
        # Aggiungi il pulsante alla lista dei widget per pulizia
        self.edit_widgets.append(calendar_btn)
        
        return date_frame, entry_widget

    def create_widgets(self):
        main_frame = tk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        left_frame = tk.Frame(main_frame, width=200)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        
        self.content_frame = tk.Frame(main_frame)
        self.content_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Sezione File XML - Include il pulsante "Modifica Fattura"
        xml_section = tk.LabelFrame(left_frame, text="File XML", padx=5, pady=5)
        xml_section.pack(fill=tk.X, padx=0, pady=(0, 10))
        
        xml_btn = tk.Button(xml_section, text="Seleziona Fattura XML", command=self.select_xml, width=20)
        xml_btn.pack(anchor=tk.W, pady=(0, 5))
        
        template_btn = tk.Button(xml_section, text="Carica Modello", command=self.load_template, 
                            bg="#8BC34A", fg="white", width=20)
        template_btn.pack(anchor=tk.W, pady=(0, 5))
        
        # Salva il riferimento al pulsante edit_btn e lo inizializza come disabilitato
        self.edit_btn = tk.Button(xml_section, text="Modifica Fattura", command=self.edit_invoice, 
                            bg="#2196F3", fg="white", width=20, state=tk.DISABLED)
        self.edit_btn.pack(anchor=tk.W, pady=(0, 5))
        
        # Sezione Foglio di stile
        xsl_section = tk.LabelFrame(left_frame, text="Foglio di stile", padx=5, pady=5)
        xsl_section.pack(fill=tk.X, padx=0, pady=(0, 10))
        
        self.xsl_var = tk.StringVar()
        self.xsl_dropdown = ttk.Combobox(xsl_section, textvariable=self.xsl_var, width=20, state="readonly")
        self.xsl_dropdown["values"] = []
        self.xsl_dropdown.pack(anchor=tk.W, pady=(0, 5))
        self.xsl_dropdown.bind("<<ComboboxSelected>>", self.on_xsl_selected)
        
        # Salva il riferimento al pulsante view_btn e lo inizializza come disabilitato
        self.view_btn = tk.Button(xsl_section, text="Visualizza Fattura", command=self.transform_and_view, 
                            bg="#4CAF50", fg="white", width=20, state=tk.DISABLED)
        self.view_btn.pack(anchor=tk.W, pady=(5, 5))
        
        # Sezione Excel - Aggiornata con quattro pulsanti
        excel_section = tk.LabelFrame(left_frame, text="Gestione Excel", padx=5, pady=5)
        excel_section.pack(fill=tk.X, padx=0, pady=(0, 10))
            
        # Sezione Excel - Aggiungere l'etichetta per il database
        excel_section = tk.LabelFrame(left_frame, text="Gestione Excel", padx=5, pady=5)
        excel_section.pack(fill=tk.X, padx=0, pady=(0, 10))
        
        # Pulsante per caricare un database Excel esistente
        excel_load_btn = tk.Button(excel_section, text="Carica DB Excel", command=self.load_excel_db,
                                bg="#673AB7", fg="white", width=20)
        excel_load_btn.pack(anchor=tk.W, pady=(0, 5))
        
        # Pulsante per creare un nuovo database Excel
        excel_create_btn = tk.Button(excel_section, text="Crea DB Excel", command=self.create_excel_db,
                                bg="#009688", fg="white", width=20)
        excel_create_btn.pack(anchor=tk.W, pady=(0, 5))
        
        # Etichetta per mostrare il database Excel attuale
        db_label_frame = tk.Frame(excel_section)
        db_label_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(db_label_frame, text="Database:").pack(side=tk.LEFT, padx=(0, 5))
        self.excel_db_label = tk.Label(db_label_frame, text="Non specificato", fg="gray", anchor="w")
        self.excel_db_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Salva il riferimento al pulsante excel_save_btn e lo inizializza come disabilitato
        self.excel_save_btn = tk.Button(excel_section, text="Salva in Excel", command=self.export_to_excel,
                                bg="#2196F3", fg="white", width=20, state=tk.DISABLED)
        self.excel_save_btn.pack(anchor=tk.W, pady=(0, 5))
        
        # Salva il riferimento al pulsante excel_manage_btn e lo inizializza come disabilitato
        self.excel_manage_btn = tk.Button(excel_section, text="Gestisci Fatture", command=self.manage_invoices,
                                bg="#FF9800", fg="white", width=20, state=tk.DISABLED)
        self.excel_manage_btn.pack(anchor=tk.W, pady=(0, 5))
        
        # Informazioni sui file
        info_frame = tk.LabelFrame(self.content_frame, text="Informazioni sui file")
        info_frame.pack(fill=tk.X, padx=5, pady=5)
        
        tk.Label(info_frame, text="File XML:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.xml_label = tk.Label(info_frame, text="Nessun file selezionato")
        self.xml_label.grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)
        
        tk.Label(info_frame, text="File XSL:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.xsl_label = tk.Label(info_frame, text="Nessun file selezionato")
        self.xsl_label.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        
        self.xsl_full_path_label = tk.Label(info_frame, text="", font=("", 8), fg="gray")
        self.xsl_full_path_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=5, pady=2)
        
        self.log_frame = tk.LabelFrame(self.content_frame, text="Log")
        self.log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(self.log_frame, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.editor_frame = tk.LabelFrame(self.content_frame, text="Modifica Fattura")
        # L'editor viene mostrato solo in modalità modifica
        
        self.editor_canvas = tk.Canvas(self.editor_frame)
        editor_scrollbar = ttk.Scrollbar(self.editor_frame, orient="vertical", command=self.editor_canvas.yview)
        self.editor_scrollable_frame = ttk.Frame(self.editor_canvas)
        
        self.editor_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.editor_canvas.configure(scrollregion=self.editor_canvas.bbox("all"))
        )
        
        self.editor_canvas.create_window((0, 0), window=self.editor_scrollable_frame, anchor="nw")
        self.editor_canvas.configure(yscrollcommand=editor_scrollbar.set)
        
        self.editor_canvas.pack(side="left", fill="both", expand=True)
        editor_scrollbar.pack(side="right", fill="y")
        
        self.editor_buttons_frame = tk.Frame(self.editor_frame)
        self.editor_buttons_frame.pack(fill=tk.X, pady=5)
        
        self.save_btn = tk.Button(self.editor_buttons_frame, text="Salva XML", command=self.save_xml,
                                bg="#FFC107", fg="black", padx=10, pady=5)
        self.save_btn.pack(side=tk.RIGHT, padx=5)
        
        self.cancel_btn = tk.Button(self.editor_buttons_frame, text="Annulla", command=self.cancel_edit,
                                    padx=10, pady=5)
        self.cancel_btn.pack(side=tk.RIGHT, padx=5)
        
        self.log("Applicazione avviata. Seleziona i file per iniziare.")

        
            
    def find_xsl_files(self):
        try:
            xsl_pattern = os.path.join(self.project_dir, "*.xsl")
            self.xsl_files = glob.glob(xsl_pattern)
            
            xsl_dir = os.path.join(self.project_dir, "xsl")
            if os.path.exists(xsl_dir) and os.path.isdir(xsl_dir):
                xsl_pattern = os.path.join(xsl_dir, "*.xsl")
                self.xsl_files.extend(glob.glob(xsl_pattern))
            
            if self.xsl_files:
                self.xsl_dropdown["values"] = [os.path.basename(f) for f in self.xsl_files]
                self.xsl_dropdown.current(0)
                self.xsl_path = self.xsl_files[0]
                self.update_xsl_labels(self.xsl_files[0])
                self.log(f"Trovati {len(self.xsl_files)} fogli di stile XSL")
        except Exception as e:
            self.log(f"Errore durante la ricerca dei file XSL: {str(e)}")
    
    # Modifica le funzioni che caricano file per aggiornare lo stato dei pulsanti
    def select_xml(self):
        filepath = filedialog.askopenfilename(
            title="Seleziona il file XML della fattura",
            filetypes=[("File XML", "*.xml")]
        )
        if filepath:
            self.xml_path = filepath
            self.xml_label.config(text=os.path.basename(filepath))
            self.log(f"File XML selezionato: {filepath}")
            try:
                self.xml_doc = etree.parse(filepath)
                self.log("File XML caricato con successo")
                # Aggiorna lo stato dei pulsanti
                self.update_button_states()
            except Exception as e:
                self.log(f"Errore nel caricamento del file XML: {str(e)}")
                self.xml_doc = None
                # Aggiorna lo stato dei pulsanti (disabilita)
                self.update_button_states()
    
    def on_xsl_selected(self, event):
        selected_index = self.xsl_dropdown.current()
        if selected_index >= 0 and selected_index < len(self.xsl_files):
            self.xsl_path = self.xsl_files[selected_index]
            self.update_xsl_labels(self.xsl_path)
            self.log(f"Foglio di stile selezionato: {self.xsl_path}")
            # Aggiorna lo stato dei pulsanti
            self.update_button_states()
    
    def update_xsl_labels(self, filepath):
        self.xsl_label.config(text=os.path.basename(filepath))
        self.xsl_full_path_label.config(text=filepath)
    
    def transform_and_view(self):
        if not self.xml_path:
            messagebox.showerror("Errore", "Seleziona il file XML della fattura")
            return
        if not self.xsl_path:
            messagebox.showerror("Errore", "Seleziona un foglio di stile XSL")
            return
        try:
            xml_doc = etree.parse(self.xml_path)
            xsl_doc = etree.parse(self.xsl_path)
            transformer = etree.XSLT(xsl_doc)
            result = transformer(xml_doc)
            fd, temp_path = tempfile.mkstemp(suffix='.html')
            with os.fdopen(fd, 'wb') as f:
                f.write(etree.tostring(result, pretty_print=True))
            webbrowser.open('file://' + temp_path)
            self.log("Trasformazione completata. Visualizzazione nel browser.")
        except Exception as e:
            self.log(f"Errore durante la trasformazione: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore durante la trasformazione:\n{str(e)}")
            
    def edit_invoice(self):
        if not self.xml_path or not self.xml_doc:
            messagebox.showerror("Errore", "Seleziona prima un file XML valido")
            return
        
        for widget in self.edit_widgets:
            if widget.winfo_exists():
                widget.destroy()
        self.edit_widgets = []
        self.line_modifications = {}
        
        self.log_frame.pack_forget()
        
        # Riconfigura il frame dell'editor per una migliore visualizzazione
        self.editor_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Rimuovi i vecchi pulsanti dell'editor se esistono
        if hasattr(self, 'editor_buttons_frame') and self.editor_buttons_frame.winfo_exists():
            self.editor_buttons_frame.destroy()
        
        # Riconfigura il canvas e la scrollbar
        self.editor_canvas.delete("all")
        self.editor_scrollable_frame = ttk.Frame(self.editor_canvas)
        self.editor_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.editor_canvas.configure(scrollregion=self.editor_canvas.bbox("all"))
        )
        self.editor_canvas.create_window((0, 0), window=self.editor_scrollable_frame, anchor="nw")
        
        # Funzione per gestire lo scrolling con rotellina del mouse da qualunque widget figlio
        def _on_mousewheel(event):
            self.editor_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        # Aggiungi la gestione scrolling al frame principale e a tutti i sottowdiget
        self.editor_frame.bind_all("<MouseWheel>", _on_mousewheel)
        
        self.log("Scansione documento XML...")
        try:
            root = self.xml_doc.getroot()
            num_elements = len(root.xpath("//*", namespaces=self.NS))
            self.log(f"Documento XML: {self.xml_path}")
            self.log(f"Numero di elementi: {num_elements}")
            self.log(f"Namespace utilizzato: {self.NS}")
        except Exception as e:
            self.log(f"Errore nell'analisi del documento XML: {str(e)}")
        
        # Crea i campi di modifica con il nuovo layout
        self.create_edit_fields()
        
        # Aggiorna i totali nel riepilogo
        self.update_riepilogo_totals()        
        self.log("Modalità modifica attivata")


    def try_find_element(self, path, namespaces):
        try:
            elements = self.xml_doc.getroot().xpath(path, namespaces=namespaces)
            if elements and len(elements) > 0:
                element = elements[0]
                return element, element.text or ""
            else:
                self.log(f"Elemento non trovato con il percorso: {path}")
                return None, ""
        except Exception as e:
            self.log(f"Errore nel trovare il campo {path}: {str(e)}")
            return None, ""



    def create_edit_fields(self):
        root = self.xml_doc.getroot()
        ns = self.NS
        
        # Definisci quali campi sono campi data
        date_fields = [
            "//p:FatturaElettronica/FatturaElettronicaBody/DatiGenerali/DatiGeneraliDocumento/Data",
            "//p:FatturaElettronica/FatturaElettronicaBody/DatiPagamento/DettaglioPagamento/DataScadenzaPagamento"
        ]
        
        # Valori possibili per Aliquota IVA
        aliquote_iva = ["4.00", "5.00", "10.00", "22.00"]
        
        # Definizione delle sezioni
        left_sections = [
            ("Dati Intestazione", [
                ("//p:FatturaElettronica/FatturaElettronicaHeader/DatiTrasmissione/IdTrasmittente/IdPaese", "ID Paese"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/DatiTrasmissione/IdTrasmittente/IdCodice", "ID Codice"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/DatiTrasmissione/ProgressivoInvio", "Progressivo Invio"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/DatiTrasmissione/FormatoTrasmissione", "Formato Trasmissione"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/DatiTrasmissione/CodiceDestinatario", "Codice Destinatario")
            ]),
            ("Cedente/Prestatore", [
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/DatiAnagrafici/IdFiscaleIVA/IdPaese", "ID Paese"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/DatiAnagrafici/IdFiscaleIVA/IdCodice", "Partita IVA"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/DatiAnagrafici/Anagrafica/Denominazione", "Denominazione"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/DatiAnagrafici/RegimeFiscale", "Regime Fiscale"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/Sede/Indirizzo", "Indirizzo"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/Sede/CAP", "CAP"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/Sede/Comune", "Comune"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/Sede/Provincia", "Provincia"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CedentePrestatore/Sede/Nazione", "Nazione")
            ]),
            ("Cessionario/Committente", [
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/DatiAnagrafici/IdFiscaleIVA/IdPaese", "ID Paese"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/DatiAnagrafici/IdFiscaleIVA/IdCodice", "Partita IVA"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/DatiAnagrafici/CodiceFiscale", "Codice Fiscale"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/DatiAnagrafici/Anagrafica/Denominazione", "Denominazione"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/Sede/Indirizzo", "Indirizzo"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/Sede/CAP", "CAP"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/Sede/Comune", "Comune"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/Sede/Provincia", "Provincia"),
                ("//p:FatturaElettronica/FatturaElettronicaHeader/CessionarioCommittente/Sede/Nazione", "Nazione")
            ])
        ]
        
        right_sections = [
            ("Dati Generali Documento", [
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiGenerali/DatiGeneraliDocumento/TipoDocumento", "Tipo Documento"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiGenerali/DatiGeneraliDocumento/Divisa", "Divisa"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiGenerali/DatiGeneraliDocumento/Data", "Data"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiGenerali/DatiGeneraliDocumento/Numero", "Numero"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiGenerali/DatiGeneraliDocumento/ImportoTotaleDocumento", "Importo Totale")
            ]),
            ("Dati Riepilogo", [
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi/DatiRiepilogo/AliquotaIVA", "Aliquota IVA"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi/DatiRiepilogo/ImponibileImporto", "Imponibile"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi/DatiRiepilogo/Imposta", "Imposta"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi/DatiRiepilogo/EsigibilitaIVA", "Esigibilità IVA")
            ]),
            ("Dati Pagamento", [
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiPagamento/CondizioniPagamento", "Condizioni Pagamento"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiPagamento/DettaglioPagamento/ModalitaPagamento", "Modalità Pagamento"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiPagamento/DettaglioPagamento/DataScadenzaPagamento", "Data Scadenza"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiPagamento/DettaglioPagamento/ImportoPagamento", "Importo Pagamento"),
                ("//p:FatturaElettronica/FatturaElettronicaBody/DatiPagamento/DettaglioPagamento/CodicePagamento", "Codice Pagamento")
            ])
        ]
        
        # Creazione del layout a due colonne
        columns_frame = tk.Frame(self.editor_scrollable_frame)
        columns_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=5)
        self.edit_widgets.append(columns_frame)
        
        # Colonna sinistra
        left_column = tk.Frame(columns_frame)
        left_column.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Colonna destra
        right_column = tk.Frame(columns_frame)
        right_column.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # Inizializza i dizionari e le liste
        self.edit_widgets = [columns_frame, left_column, right_column]
        self.edit_fields = {}
        
        # Dizionario per tenere traccia delle variabili StringVar dei campi
        self.string_vars = {}
        
        # Popola la colonna sinistra
        for section_title, fields in left_sections:
            section_frame = tk.LabelFrame(left_column, text=section_title)
            section_frame.pack(fill=tk.BOTH, expand=True, pady=5)
            self.edit_widgets.append(section_frame)
            
            # Variabili per tenere traccia dei widget di comune, provincia e CAP
            comune_var = None
            provincia_var = None
            cap_var = None
            comune_widget = None
            provincia_xpath = None
            cap_xpath = None
            comune_xpath = None
            comune_row = None
            
            for i, (xpath, label) in enumerate(fields):
                element, value = self.try_find_element(xpath, self.NS)
                label_widget = tk.Label(section_frame, text=label + ":")
                label_widget.grid(row=i, column=0, sticky="w", padx=5, pady=2)
                
                # Crea una variabile StringVar per ogni campo
                string_var = tk.StringVar(value=value)
                self.string_vars[xpath] = string_var
                
                # Verifica se è un campo data
                if xpath in date_fields:
                    # Usa il metodo create_date_field per creare il campo data
                    date_frame, entry_widget = self.create_date_field(section_frame, value, width=28)
                    date_frame.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                else:
                    # Trattamento speciale per i campi Comune, Provincia e CAP
                    if "Comune" in label and section_title in ["Cedente/Prestatore", "Cessionario/Committente"]:
                        # Salva il riferimento alla variabile e xpath del comune
                        comune_var = string_var
                        comune_xpath = xpath
                        comune_row = i
                        # Crea un Entry normale per ora, sarà sostituito dopo
                        entry_widget = tk.Entry(section_frame, width=30, textvariable=string_var)
                        entry_widget.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                        comune_widget = entry_widget
                    elif "Provincia" in label and section_title in ["Cedente/Prestatore", "Cessionario/Committente"]:
                        provincia_var = string_var
                        provincia_xpath = xpath
                        entry_widget = tk.Entry(section_frame, width=30, textvariable=string_var)
                        entry_widget.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                    elif "CAP" in label and section_title in ["Cedente/Prestatore", "Cessionario/Committente"]:
                        cap_var = string_var
                        cap_xpath = xpath
                        entry_widget = tk.Entry(section_frame, width=30, textvariable=string_var)
                        entry_widget.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                    else:
                        # Per i campi normali, usa Entry standard con StringVar
                        entry_widget = tk.Entry(section_frame, width=30, textvariable=string_var)
                        entry_widget.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                
                self.edit_fields[xpath] = {"widget": entry_widget, "element": element}
                self.edit_widgets.extend([label_widget, entry_widget])
                
                # Configura ridimensionamento
                section_frame.columnconfigure(1, weight=1)
            
            # Dopo aver creato tutti i campi nella sezione, se abbiamo trovato i campi necessari
            # per l'autocompletamento, sostituisci il widget del comune
            if comune_var and provincia_var and cap_var and comune_widget and comune_row is not None:
                # Rimuovi il vecchio widget
                comune_widget.destroy()
                
                # Crea il nuovo widget di autocompletamento
                autocomplete = AutocompleteComune(section_frame, comune_var, provincia_var, cap_var, width=28)
                autocomplete.frame.grid(row=comune_row, column=1, sticky="ew", padx=5, pady=2)
                
                # Aggiorna il riferimento nel dizionario edit_fields
                if comune_xpath:
                    # Aggiorna il riferimento al widget
                    self.edit_fields[comune_xpath]["widget"] = autocomplete.comune_entry
                
                # Aggiungi alla lista dei widget da distruggere
                self.edit_widgets.append(autocomplete.frame)
                
                # Log per debug
                self.log(f"Abilitato autocompletamento comuni per la sezione {section_title}")
        
        # Popola la colonna destra
        for section_title, fields in right_sections:
            section_frame = tk.LabelFrame(right_column, text=section_title)
            section_frame.pack(fill=tk.BOTH, expand=True, pady=5)
            self.edit_widgets.append(section_frame)
            
            # Variabili di riferimento per i campi dati riepilogo
            self.aliquota_iva_widget = None
            self.imponibile_widget = None
            self.imposta_widget = None
            
            for i, (xpath, label) in enumerate(fields):
                element, value = self.try_find_element(xpath, self.NS)
                label_widget = tk.Label(section_frame, text=label + ":")
                label_widget.grid(row=i, column=0, sticky="w", padx=5, pady=2)
                
                # Per AliquotaIVA creiamo un Combobox invece di un Entry
                if "AliquotaIVA" in xpath:
                    entry_widget = ttk.Combobox(section_frame, width=29, state="readonly")
                    entry_widget['values'] = aliquote_iva
                    
                    # Trova il valore corrispondente nella lista o seleziona quello più vicino
                    if value in aliquote_iva:
                        entry_widget.set(value)
                    else:
                        # Se non troviamo un match esatto, cerchiamo il più vicino
                        try:
                            if value:
                                val_float = float(value)
                                closest = min(aliquote_iva, key=lambda x: abs(float(x) - val_float))
                                entry_widget.set(closest)
                            else:
                                # Default a 22% se non c'è un valore
                                entry_widget.set("22.00")
                        except:
                            # In caso di errore, default a 22%
                            entry_widget.set("22.00")                 
                    
                    entry_widget.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                    
                    # Se siamo nella sezione Dati Riepilogo, tieni traccia del widget
                    if section_title == "Dati Riepilogo":
                        self.aliquota_iva_widget = entry_widget
                    
                # Verifica se è un campo data
                elif xpath in date_fields:
                    # Usa il metodo create_date_field per creare il campo data
                    date_frame, entry_widget = self.create_date_field(section_frame, value, width=28)
                    date_frame.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                else:
                    # Per i campi normali, usa Entry standard
                    entry_widget = tk.Entry(section_frame, width=30)
                    entry_widget.insert(0, value)
                    entry_widget.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
                    
                    # Tieni traccia dei campi imponibile e imposta nella sezione Dati Riepilogo
                    if section_title == "Dati Riepilogo":
                        if "ImponibileImporto" in xpath:
                            self.imponibile_widget = entry_widget
                        elif "Imposta" in xpath:
                            self.imposta_widget = entry_widget
                
                self.edit_fields[xpath] = {"widget": entry_widget, "element": element}
                self.edit_widgets.extend([label_widget, entry_widget])
                
                # Configura ridimensionamento
                section_frame.columnconfigure(1, weight=1)
                
            # Se siamo nella sezione Dati Riepilogo e abbiamo i riferimenti necessari,
            # configura il calcolo automatico dell'imposta
            if section_title == "Dati Riepilogo" and self.aliquota_iva_widget and self.imponibile_widget and self.imposta_widget:
                # Crea una variabile per tracciare le modifiche
                self.aliquota_var = tk.StringVar()
                self.aliquota_var.set(self.aliquota_iva_widget.get())  # Valore iniziale
                self.aliquota_iva_widget.config(textvariable=self.aliquota_var)
                
                # Doppio binding: sia all'evento che alla variabile
                self.aliquota_iva_widget.bind("<<ComboboxSelected>>", self.calcola_imposta)
                self.aliquota_var.trace_add("write", lambda *args: self.calcola_imposta())
                
                # Evento per l'imponibile quando perde il focus
                self.imponibile_widget.bind("<FocusOut>", self.calcola_imposta)
                
                # Evento per l'imponibile quando viene premuto Invio
                self.imponibile_widget.bind("<Return>", self.calcola_imposta)
                
                # Colora di grigio chiaro il campo imposta per indicare che è calcolato automaticamente
                self.imposta_widget.config(bg="#f0f0f0")
                
                # Aggiungi un tooltip visibile
                imposta_label = [lbl for lbl in section_frame.winfo_children() 
                                if isinstance(lbl, tk.Label) and "Imposta" in lbl["text"]][0]
                imposta_label.config(text="Imposta (auto):")
                
                # Esegui il calcolo iniziale dopo un breve ritardo
                self.after(800, self.calcola_imposta)
        
        # Linee di dettaglio come riga estesa sotto entrambe le colonne
        detail_row = 1
        
        # Aggiorna i dati delle linee
        self.refresh_lines_data()
        
        # Gestione dei campi per DettaglioLinee
        self.line_row = detail_row
        self.current_line_index = 0
        
        if not hasattr(self, 'total_lines') or self.total_lines == 0:
            self.total_lines = len(self.normal_lines)
        
        self.log(f"Trovate {self.total_lines} linee di dettaglio normali")
        
        # Frame per linee di dettaglio che si estende a larghezza completa
        details_full_width = tk.Frame(self.editor_scrollable_frame)
        details_full_width.grid(row=detail_row, column=0, sticky="ew", padx=10, pady=5)
        self.edit_widgets.append(details_full_width)
        
        # Crea i campi per la linea di dettaglio corrente
        self.update_line_fields(details_full_width)
        
        # Checkbox per la linea CONAI
        conai_frame = tk.Frame(self.editor_scrollable_frame)
        conai_frame.grid(row=detail_row+1, column=0, sticky="ew", padx=10, pady=5)
        self.edit_widgets.append(conai_frame)

        self.conai_var = tk.BooleanVar()
        self.conai_var.set(self.conai_line is not None)

        conai_check = tk.Checkbutton(conai_frame, text="Includi linea CONTRIBUTO CONAI ASSOLTO", 
                                    variable=self.conai_var, command=self.toggle_conai,
                                    padx=5, pady=5, font=("", 10, "bold"))
        conai_check.pack(side=tk.LEFT)
        self.edit_widgets.append(conai_check)

        # Frame di navigazione
        nav_frame = tk.Frame(self.editor_scrollable_frame)
        nav_frame.grid(row=detail_row+2, column=0, sticky="ew", padx=10)
        self.edit_widgets.append(nav_frame)
        
        def prev_line():
            if self.current_line_index > 0:
                # Memorizza la posizione attuale dello scrolling
                current_scroll_position = self.editor_canvas.yview()
                
                self.save_current_line_data()
                self.current_line_index -= 1
                self.update_line_fields(details_full_width)
                self.update_nav_buttons()
                
                # Ripristina la posizione dello scrolling dopo l'aggiornamento
                self.editor_canvas.update_idletasks()  # Assicura che tutti gli aggiornamenti del layout siano completi
                self.editor_canvas.yview_moveto(current_scroll_position[0])
        
        def next_line():
            if self.current_line_index < self.total_lines - 1:
                # Memorizza la posizione attuale dello scrolling
                current_scroll_position = self.editor_canvas.yview()
                
                self.save_current_line_data()
                self.current_line_index += 1
                self.update_line_fields(details_full_width)
                self.update_nav_buttons()
                
                # Ripristina la posizione dello scrolling dopo l'aggiornamento
                self.editor_canvas.update_idletasks()  # Assicura che tutti gli aggiornamenti del layout siano completi
                self.editor_canvas.yview_moveto(current_scroll_position[0])

        self.prev_btn = tk.Button(nav_frame, text="◀ Prec", command=prev_line)
        self.prev_btn.pack(side=tk.LEFT, padx=5)
        
        self.next_btn = tk.Button(nav_frame, text="Succ ▶", command=next_line)
        self.next_btn.pack(side=tk.LEFT, padx=5)
        
        self.line_label = tk.Label(nav_frame, text=f"Linea {self.current_line_index + 1} di {self.total_lines}")
        self.line_label.pack(side=tk.LEFT, padx=10)
        
        add_btn = tk.Button(nav_frame, text="➕ Aggiungi", command=self.add_line, bg="#4CAF50", fg="white")
        add_btn.pack(side=tk.RIGHT, padx=5)
        
        del_btn = tk.Button(nav_frame, text="➖ Elimina", command=self.delete_line, bg="#F44336", fg="white")
        del_btn.pack(side=tk.RIGHT, padx=5)

        recalc_btn = tk.Button(nav_frame, text="🔄 Ricalcola Totali", 
                            command=self.update_riepilogo_totals, 
                            bg="#FFA500", fg="white")
        recalc_btn.pack(side=tk.RIGHT, padx=5)
        self.edit_widgets.append(recalc_btn)
                
        self.update_nav_buttons()
        
        # Frame per i pulsanti centrati in basso
        buttons_frame = tk.Frame(self.editor_scrollable_frame)
        buttons_frame.grid(row=detail_row+3, column=0, sticky="ew", padx=10, pady=15)
        self.edit_widgets.append(buttons_frame)
        
        # Contenitore per centrare i pulsanti
        centered_buttons = tk.Frame(buttons_frame)
        centered_buttons.pack(fill=tk.X)
        self.edit_widgets.append(centered_buttons)
        
        # Spazio vuoto a sinistra per centrare
        left_spacer = tk.Frame(centered_buttons)
        left_spacer.pack(side=tk.LEFT, expand=True)
        
        # Pulsanti centrati
        self.cancel_btn = tk.Button(centered_buttons, text="Annulla", command=self.cancel_edit,
                                    padx=20, pady=8, font=("", 10))
        self.cancel_btn.pack(side=tk.LEFT, padx=10)
        
        self.save_btn = tk.Button(centered_buttons, text="Salva XML", command=self.save_xml,
                                bg="#FFC107", fg="black", padx=20, pady=8, font=("", 10, "bold"))
        self.save_btn.pack(side=tk.LEFT, padx=10)
        
        # Spazio vuoto a destra per centrare
        right_spacer = tk.Frame(centered_buttons)
        right_spacer.pack(side=tk.RIGHT, expand=True)
        
        # Pulsante per visualizzare la struttura XML
        tree_btn = tk.Button(buttons_frame, text="Visualizza struttura XML completa", 
                            command=self.show_xml_tree, bg="#607D8B", fg="white", padx=10, pady=5)
        tree_btn.pack(side=tk.BOTTOM, pady=(10, 0))
        self.edit_widgets.append(tree_btn)
        
        # Aggiorna i totali nel riepilogo
        self.update_riepilogo_totals()
        
        self.editor_scrollable_frame.update_idletasks()
        self.editor_canvas.config(width=self.editor_scrollable_frame.winfo_reqwidth())


        
    def on_line_field_change(self, event, xpath):
        if self.current_line_index not in self.line_modifications:
            self.line_modifications[self.current_line_index] = {}
        self.line_modifications[self.current_line_index][xpath] = event.widget.get()
    
    def show_xml_tree(self):
        if not self.xml_doc:
            messagebox.showerror("Errore", "Nessun documento XML caricato")
            return
        tree_window = tk.Toplevel(self)
        tree_window.title("Struttura XML")
        tree_window.geometry("800x600")
        frame = tk.Frame(tree_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget = tk.Text(frame, wrap=tk.NONE, yscrollcommand=scrollbar.set)
        text_widget.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        xml_string = etree.tostring(self.xml_doc, pretty_print=True, encoding='utf-8').decode('utf-8')
        text_widget.insert(tk.END, xml_string)
        text_widget.config(state=tk.DISABLED)
        
    def save_xml(self):
        if not self.xml_doc:
            return

        if hasattr(self, 'save_current_line_data'):
            self.save_current_line_data()
        
        modifiche_effettuate = []
        for xpath, field_data in list(self.edit_fields.items()):
            widget = field_data["widget"]
            element = field_data["element"]
            try:
                new_value = widget.get()
            except (tk.TclError, AttributeError):
                continue
            if element is not None:
                old_value = element.text or ""
                if old_value != new_value:
                    element.text = new_value
                    element_name = element.tag.split('}')[-1]
                    modifiche_effettuate.append((element_name, old_value, new_value))
        
        for line_index, line_data in self.line_modifications.items():
            for xpath, new_value in line_data.items():
                elements = self.xml_doc.getroot().xpath(xpath, namespaces=self.NS)
                if elements:
                    elements[0].text = new_value
        
        output_path = filedialog.asksaveasfilename(
            title="Salva XML modificato",
            defaultextension=".xml",
            filetypes=[("File XML", "*.xml")],
            initialfile=os.path.basename(self.xml_path)
        )
        
        if output_path:
            try:
                self.indent(self.xml_doc.getroot())
                new_xml = etree.tostring(self.xml_doc, pretty_print=True, encoding="UTF-8", xml_declaration=True).decode("utf-8")
                new_xml = re.sub(r'(</DettaglioLinee>)(\r?\n)+(<(?:\w+:)?DatiRiepilogo>)', r'\1\n      \3', new_xml)
                new_xml = re.sub(r'(</DettaglioLinee>)(<(?:\w+:)?DatiRiepilogo>)', r'\1\n      \2', new_xml)
                
                # Aggiungi il prefisso "p:" SOLO all'elemento radice FatturaElettronica
                new_xml = new_xml.replace("<FatturaElettronica ", "<p:FatturaElettronica ")
                new_xml = new_xml.replace("</FatturaElettronica>", "</p:FatturaElettronica>")
                
                # Aggiungi il namespace "xmlns:p" all'elemento radice
                new_xml = new_xml.replace("<p:FatturaElettronica ", 
                                        "<p:FatturaElettronica xmlns:p=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\" ")
                
                # Rimuovi il namespace senza prefisso, per evitare di avere entrambi
                new_xml = new_xml.replace(" xmlns=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\"", "")
                
                # Aggiungi il riferimento allo stylesheet XSL
                stylesheet_ref = '<?xml-stylesheet type="text/xsl" href="./fatturapa_v1.2_asw.xsl"?>\n'
                if '<?xml ' in new_xml:
                    xml_decl_end = new_xml.find('?>') + 2
                    new_xml = new_xml[:xml_decl_end] + '\n' + stylesheet_ref + new_xml[xml_decl_end:]
                else:
                    new_xml = stylesheet_ref + new_xml
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(new_xml)
                
                # Aggiorna il percorso del file XML
                self.xml_path = output_path
                self.xml_label.config(text=os.path.basename(output_path))
                
                self.log(f"File salvato con successo: {output_path}")
                messagebox.showinfo("Salvataggio completato", f"Il file XML è stato salvato con successo.")
                
                # Verifica se il file salvato era un file temporaneo
                # e chiedi all'utente se vuole aggiornare il database Excel
                is_temp_file = "temp_invoice_" in os.path.basename(self.xml_path).lower()
                
                if is_temp_file and self.excel_manager.excel_path:
                    update_excel = messagebox.askyesno("Aggiornamento Excel", 
                                                    "Vuoi aggiornare anche il database Excel con le modifiche apportate?")
                    if update_excel:
                        success = self.excel_manager.export_xml_to_excel(self.xml_doc)
                        if success:
                            messagebox.showinfo("Aggiornamento Excel", 
                                            "Il database Excel è stato aggiornato con successo.")
                        else:
                            messagebox.showerror("Errore", 
                                            "Si è verificato un errore durante l'aggiornamento del database Excel.")
                
                self.cancel_edit()
                
                # Aggiorna lo stato dei pulsanti
                self.update_button_states()
            except Exception as e:
                self.log(f"Errore nel salvataggio del file: {str(e)}")
                messagebox.showerror("Errore", f"Errore nel salvataggio del file:\n{str(e)}")
  
    def cancel_edit(self):
        # Rimuovi il binding della rotellina del mouse
        self.unbind_all("<MouseWheel>")
        
        self.editor_frame.pack_forget()
        self.log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log("Modalità modifica disattivata")
    
    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
    


    def add_conai_line(self):
        try:
            # Verifica se la linea CONAI esiste già
            self.refresh_lines_data()
            if self.conai_line is not None:
                self.log("La linea CONTRIBUTO CONAI ASSOLTO è già presente")
                return
            
            # Ottieni il nodo DatiBeniServizi
            dati_beni = self.xml_doc.getroot().xpath("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi", namespaces=self.NS)[0]
            
            # Crea la linea CONAI
            conai_line = etree.Element("DettaglioLinee")
            etree.SubElement(conai_line, "NumeroLinea").text = "0"  # Sarà aggiornato dopo
            etree.SubElement(conai_line, "Descrizione").text = "CONTRIBUTO CONAI ASSOLTO"
            etree.SubElement(conai_line, "PrezzoUnitario").text = "0.0000000"
            etree.SubElement(conai_line, "PrezzoTotale").text = "0.0000000"
            etree.SubElement(conai_line, "AliquotaIVA").text = "22.00"
            
            # Inserisci prima di DatiRiepilogo se esiste
            riepilogo = dati_beni.xpath("./DatiRiepilogo", namespaces=self.NS)
            if riepilogo:
                dati_beni.insert(dati_beni.index(riepilogo[0]), conai_line)
            else:
                dati_beni.append(conai_line)
            
            # Aggiorna i numeri delle linee
            self.update_line_numbers()
            
            # Aggiorna i dati e l'interfaccia
            self.refresh_lines_data()
            self.update_line_fields()
            self.update_nav_buttons()
            
            self.log("Aggiunta linea CONTRIBUTO CONAI ASSOLTO")
        except Exception as e:
            self.log(f"Errore nell'aggiunta della linea CONAI: {str(e)}")
            traceback.print_exc()
            
    def remove_conai_line(self):
        try:
            # Verifica se la linea CONAI esiste
            self.refresh_lines_data()
            if self.conai_line is None:
                self.log("Nessuna linea CONTRIBUTO CONAI ASSOLTO trovata")
                return
            
            # Rimuovi la linea CONAI
            parent = self.conai_line.getparent()
            parent.remove(self.conai_line)
            
            # Aggiorna i numeri delle linee
            self.update_line_numbers()
            
            # Aggiorna i dati e l'interfaccia
            self.refresh_lines_data()
            self.update_line_fields()
            self.update_nav_buttons()
            
            self.log("Rimossa linea CONTRIBUTO CONAI ASSOLTO")
        except Exception as e:
            self.log(f"Errore nella rimozione della linea CONAI: {str(e)}")
            traceback.print_exc()
        self.update_riepilogo_totals()

    def update_nav_buttons(self):
        self.prev_btn["state"] = "normal" if self.current_line_index > 0 else "disabled"
        self.next_btn["state"] = "normal" if self.current_line_index < self.total_lines - 1 else "disabled"
        self.line_label.config(text=f"Linea {self.current_line_index + 1} di {self.total_lines}")


    def update_line_fields(self, parent_frame=None):
        # Se parent_frame è None, usa self.editor_scrollable_frame
        if parent_frame is None:
            parent_frame = self.editor_scrollable_frame
        
        # Determina il gestore di geometria da utilizzare
        # Se parent_frame è nel editor_scrollable_frame, usa grid, altrimenti pack
        using_grid = parent_frame == self.editor_scrollable_frame
        
        # Rimuovi il vecchio frame se esiste
        if hasattr(self, 'line_frame') and self.line_frame in self.edit_widgets:
            self.line_frame.destroy()
            self.edit_widgets.remove(self.line_frame)
        
        # Rimuovi vecchi campi delle linee di dettaglio
        for key in list(self.edit_fields.keys()):
            if "DettaglioLinee" in key:
                del self.edit_fields[key]
        
        # Aggiorna i dati delle linee dal documento XML
        self.refresh_lines_data()
        
        # Crea il nuovo frame
        self.line_frame = tk.LabelFrame(parent_frame, 
                                    text=f"Dettaglio Linea {self.current_line_index + 1} di {self.total_lines}")
        
        # Usa il gestore di geometria appropriato
        if using_grid:
            self.line_frame.grid(row=self.line_row, column=0, sticky="ew", padx=10, pady=5)
        else:
            self.line_frame.pack(fill=tk.BOTH, expand=True, pady=5)
            
        self.edit_widgets.append(self.line_frame)
        
        # Se non ci sono linee normali, mostra un messaggio
        if self.total_lines == 0:
            tk.Label(self.line_frame, text="Nessuna linea di dettaglio presente").grid(row=0, column=0)
            return
        
        # Ottieni la linea corrente (solo dalle linee normali)
        current_line = self.normal_lines[self.current_line_index]
        
        # Definisci i campi da mostrare
        fields = [
            ("NumeroLinea", "Numero Linea"),
            ("Descrizione", "Descrizione"),
            ("Quantita", "Quantità"),
            ("UnitaMisura", "Unità Misura"),
            ("PrezzoUnitario", "Prezzo Unitario"),
            ("PrezzoTotale", "Prezzo Totale"),
            ("AliquotaIVA", "Aliquota IVA")
        ]
        
        # Valori possibili per Aliquota IVA
        aliquote_iva = ["4.00", "5.00", "10.00", "22.00"]
        
        # Crea un frame a griglia per i campi di dettaglio per una migliore organizzazione
        line_grid = tk.Frame(self.line_frame)
        line_grid.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.edit_widgets.append(line_grid)
        
        # Variabili per tenere traccia dei widget specifici
        quantita_widget = None
        prezzo_unitario_widget = None
        prezzo_totale_widget = None
        
        # Funzione di validazione per campi numerici
        def validate_numeric_field(field_name, text):
            # Permette stringa vuota (per cancellare tutto)
            if text == "":
                return True
            
            # Permette solo numeri e un singolo punto
            if not all(c.isdigit() or c == '.' for c in text):
                return False
            
            # Controlla che ci sia al massimo un punto
            if text.count('.') > 1:
                return False
            
            # Se contiene un punto, controlla le regole per la parte decimale
            if '.' in text:
                integer_part, decimal_part = text.split('.')
                
                # Per Quantita e PrezzoUnitario, limita la parte intera a 15 cifre
                if field_name != "PrezzoTotale" and len(integer_part) > 15:
                    return False
                    
                # Non blocca la digitazione in corso nella parte decimale
                if len(decimal_part) > 7:
                    return False
            else:
                # Se non c'è il punto, controlla solo la parte intera per Quantita e PrezzoUnitario
                if field_name != "PrezzoTotale" and len(text) > 15:
                    return False
                
            return True
        
        # Funzione per formattare il campo quando perde il focus
        def format_numeric_field(event):
            widget = event.widget
            field_name = widget.field_name if hasattr(widget, 'field_name') else ""
            current_value = widget.get().replace(',', '.')
            
            try:
                # Se il campo è vuoto, imposta 0
                if not current_value:
                    formatted_value = "0.0000000"
                    widget.delete(0, tk.END)
                    widget.insert(0, formatted_value)
                    widget.config(bg="white")
                    return
                
                # Separa parte intera e decimale
                if '.' in current_value:
                    integer_part, decimal_part = current_value.split('.')
                else:
                    integer_part, decimal_part = current_value, ""
                
                # Rimuovi zeri iniziali dalla parte intera (tranne lo zero da solo)
                if integer_part != "0":
                    integer_part = integer_part.lstrip("0") or "0"
                
                # Formatta con esattamente 7 decimali
                decimal_part = (decimal_part + "0000000")[:7]
                
                # Ricostruisci il valore formattato
                formatted_value = f"{integer_part}.{decimal_part}"
                
                # Aggiorna il campo
                widget.delete(0, tk.END)
                widget.insert(0, formatted_value)
                widget.config(bg="white")  # Resetta lo sfondo
                
                # Scatena il calcolo del prezzo totale se applicabile
                if widget in [quantita_widget, prezzo_unitario_widget]:
                    calculate_total_price()
                    
            except Exception as e:
                # In caso di errore, evidenzia il campo
                widget.config(bg="#ffcccc")  # Sfondo rosso chiaro
                self.log(f"Errore nella formattazione del campo: {str(e)}")
        
        # Layout dei campi su più righe
        row = 0
        
        # Prima riga: Numero Linea
        numero_label = tk.Label(line_grid, text="Numero Linea:")
        numero_label.grid(row=row, column=0, sticky="w", padx=5, pady=5)
        
        numero_element = current_line.find("NumeroLinea")
        numero_value = numero_element.text if numero_element is not None else ""
        
        numero_entry = tk.Entry(line_grid, width=15)
        numero_entry.insert(0, numero_value)
        numero_entry.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        
        # Memorizza il riferimento al campo e all'elemento XML
        self.edit_fields[f"current_line.NumeroLinea"] = {"widget": numero_entry, "element": numero_element}
        self.edit_widgets.extend([numero_label, numero_entry])
        
        row += 1
        
        # Seconda riga: Descrizione
        descrizione_label = tk.Label(line_grid, text="Descrizione:")
        descrizione_label.grid(row=row, column=0, sticky="w", padx=5, pady=5)
        
        descrizione_element = current_line.find("Descrizione")
        descrizione_value = descrizione_element.text if descrizione_element is not None else ""
        
        descrizione_entry = tk.Entry(line_grid, width=80)
        descrizione_entry.insert(0, descrizione_value)
        descrizione_entry.grid(row=row, column=1, columnspan=5, sticky="ew", padx=5, pady=5)
        
        # Memorizza il riferimento al campo e all'elemento XML
        self.edit_fields[f"current_line.Descrizione"] = {"widget": descrizione_entry, "element": descrizione_element}
        self.edit_widgets.extend([descrizione_label, descrizione_entry])
        
        row += 1
        
        # Organizziamo gli altri campi in 2 righe di 3 colonne ciascuna
        col = 0
        for field_name, label in fields:
            # Saltiamo NumeroLinea e Descrizione che abbiamo già gestito
            if field_name in ["NumeroLinea", "Descrizione"]:
                continue
                
            label_widget = tk.Label(line_grid, text=label + ":")
            label_widget.grid(row=row, column=col*2, sticky="w", padx=5, pady=2)
            
            element = current_line.find(field_name)
            value = element.text if element is not None else ""
            
            # Per AliquotaIVA creiamo un Combobox invece di un Entry
            if field_name == "AliquotaIVA":
                entry_widget = ttk.Combobox(line_grid, width=14, state="readonly")
                entry_widget['values'] = aliquote_iva
                
                # Trova il valore corrispondente nella lista o seleziona quello più vicino
                if value in aliquote_iva:
                    entry_widget.set(value)
                else:
                    # Se non troviamo un match esatto, cerchiamo il più vicino
                    try:
                        if value:
                            val_float = float(value)
                            closest = min(aliquote_iva, key=lambda x: abs(float(x) - val_float))
                            entry_widget.set(closest)
                        else:
                            # Default a 22% se non c'è un valore
                            entry_widget.set("22.00")
                    except:
                        # In caso di errore, default a 22%
                        entry_widget.set("22.00")



            else:
                entry_widget = tk.Entry(line_grid, width=15)
                
                # Applica validazione per campi numerici
                if field_name in ["Quantita", "PrezzoUnitario", "PrezzoTotale"]:
                    # Salva il nome del campo nel widget per poterlo recuperare nel validatore
                    entry_widget.field_name = field_name
                    
                    # Crea un validatore specifico per questo campo
                    vcmd = self.register(lambda text, fname=field_name: validate_numeric_field(fname, text))
                    entry_widget.config(validate="key", validatecommand=(vcmd, '%P'))
                    entry_widget.bind("<FocusOut>", format_numeric_field)
                    
                    # Formatta il valore iniziale
                    if value:
                        try:
                            # Converte il valore esistente nel formato corretto
                            if '.' in value:
                                int_part, dec_part = value.split('.')
                                # Rimuovi zeri iniziali dalla parte intera (tranne lo zero da solo)
                                if int_part != "0":
                                    int_part = int_part.lstrip("0") or "0"
                                # Formatta con esattamente 7 decimali
                                dec_part = (dec_part + "0000000")[:7]
                                value = f"{int_part}.{dec_part}"
                            else:
                                value = f"{value.lstrip('0') or '0'}.0000000"
                        except:
                            value = "0.0000000"
                
                entry_widget.insert(0, value)
            
            entry_widget.grid(row=row, column=col*2+1, sticky="ew", padx=5, pady=2)
            
            # Memorizza il riferimento al campo e all'elemento XML
            field_key = f"current_line.{field_name}"
            self.edit_fields[field_key] = {"widget": entry_widget, "element": element}
            self.edit_widgets.extend([label_widget, entry_widget])
            
            # Salva i riferimenti ai widget specifici
            if field_name == "Quantita":
                quantita_widget = entry_widget
            elif field_name == "PrezzoUnitario":
                prezzo_unitario_widget = entry_widget
            elif field_name == "PrezzoTotale":
                prezzo_totale_widget = entry_widget
            
            # Aggiorna la posizione per il prossimo campo
            col += 1
            if col >= 3:  # 3 colonne per riga
                col = 0
                row += 1
        
        # Configura il ridimensionamento delle colonne
        for i in range(6):  # 6 colonne (3 coppie di label/entry)
            if i % 2 == 1:  # Colonne dei widget Entry
                line_grid.columnconfigure(i, weight=1)
        
        # Configura il calcolo automatico del prezzo totale
        if quantita_widget and prezzo_unitario_widget and prezzo_totale_widget:
            def calculate_total_price(*args):
                try:
                    quantita_str = quantita_widget.get().replace(',', '.')
                    prezzo_unitario_str = prezzo_unitario_widget.get().replace(',', '.')
                    
                    # Verifica che i valori non siano vuoti
                    if quantita_str and prezzo_unitario_str:
                        quantita = float(quantita_str)
                        prezzo_unitario = float(prezzo_unitario_str)
                        prezzo_totale = quantita * prezzo_unitario
                        
                        # Formatta il risultato con 7 decimali
                        prezzo_totale_formatted = f"{prezzo_totale:.7f}"
                        
                        # Aggiorna il campo prezzo totale
                        prezzo_totale_widget.delete(0, tk.END)
                        prezzo_totale_widget.insert(0, prezzo_totale_formatted)

                        # Aggiorna anche l'elemento XML corrispondente
                        for field_key, field_data in self.edit_fields.items():
                            if field_key == "current_line.PrezzoTotale" and field_data["element"] is not None:
                                field_data["element"].text = prezzo_totale_formatted
                                break
                        
                        # Aggiorna i totali nel riepilogo
                        self.after(100, self.update_riepilogo_totals)
                                                
                except (ValueError, AttributeError) as e:
                    # In caso di errore nel parsing dei numeri, non fare nulla
                    self.log(f"Errore nel calcolo del prezzo totale: {str(e)}")
            
            # Crea variabili StringVar per tracciare le modifiche
            self.quantita_var = tk.StringVar(value=quantita_widget.get())
            self.prezzo_unitario_var = tk.StringVar(value=prezzo_unitario_widget.get())
            
            # Collega le variabili ai widget
            quantita_widget.config(textvariable=self.quantita_var)
            prezzo_unitario_widget.config(textvariable=self.prezzo_unitario_var)
            
            # Aggiungi tracciatori di modifica
            self.quantita_var.trace_add("write", calculate_total_price)
            self.prezzo_unitario_var.trace_add("write", calculate_total_price)
            
            # Calcola il prezzo totale iniziale
            calculate_total_price()
            
    def save_current_line_data(self):
        line_data = {}
        
        # Assicurati che ci siano linee normali
        if self.total_lines <= 0 or self.current_line_index >= self.total_lines:
            return
        
        # Ottieni il riferimento alla linea corrente
        current_line = self.normal_lines[self.current_line_index]
        
        # Raccogli i valori dai campi di input
        for field_key, field_data in list(self.edit_fields.items()):
            if field_key.startswith("current_line."):
                field_name = field_key.split(".")[1]
                widget = field_data["widget"]
                element = field_data["element"]
                
                try:
                    value = widget.get()
                    
                    # Formatta i campi numerici prima del salvataggio
                    if field_name in ["Quantita", "PrezzoUnitario", "PrezzoTotale"]:
                        # Assicura che il valore sia nel formato corretto (15 cifre intere, 7 decimali)
                        if value:
                            if '.' in value:
                                int_part, dec_part = value.split('.')
                                # Limita la parte intera a 15 cifre
                                if field_name != "PrezzoTotale":
                                    int_part = int_part[-15:] if len(int_part) > 15 else int_part
                                # Assicura esattamente 7 decimali
                                dec_part = (dec_part + "0000000")[:7]
                                value = f"{int_part}.{dec_part}"
                            else:
                                # Se non c'è punto, aggiungi 7 zeri dopo il punto
                                value = f"{value}.0000000"
                    
                    # Aggiorna l'elemento XML
                    if element is not None:
                        element.text = value
                        self.log(f"Aggiornato campo {field_name}: {value}")
                    
                except Exception as e:
                    self.log(f"Errore nel salvataggio del campo {field_name}: {str(e)}")
        
        # Aggiorna i numeri delle linee
        self.update_line_numbers()
        
        # Aggiorna i totali nel riepilogo
        self.update_riepilogo_totals()

    def delete_line(self):
        # Aggiorna i dati delle linee per avere informazioni aggiornate
        self.refresh_lines_data()
        
        if self.total_lines <= 0:
            messagebox.showwarning("Attenzione", "Nessuna linea da eliminare")
            return
        
        try:
            # Verifica se l'indice è valido
            if self.current_line_index < 0 or self.current_line_index >= self.total_lines:
                self.log(f"Indice linea non valido: {self.current_line_index}")
                return
                
            # Ottieni la linea da eliminare (tra le linee normali)
            line_to_delete = self.normal_lines[self.current_line_index]
            
            # Elimina la linea
            parent = line_to_delete.getparent()
            parent.remove(line_to_delete)
            
            # Aggiorna i numeri delle linee
            self.update_line_numbers()
            
            # Aggiorna i dati e l'interfaccia
            self.refresh_lines_data()
            
            # Aggiusta l'indice corrente se necessario
            if self.current_line_index >= self.total_lines:
                self.current_line_index = max(0, self.total_lines - 1)
            
            self.update_line_fields()
            self.update_nav_buttons()
            
            self.log(f"Eliminata linea, ora ci sono {self.total_lines} linee normali")
        except Exception as e:
            self.log(f"Errore nell'eliminazione della linea: {str(e)}")
            traceback.print_exc()
            messagebox.showerror("Errore", f"Impossibile eliminare la linea:\n{str(e)}")
        self.update_riepilogo_totals()
            
    def add_line(self):
        try:
            # Ottieni il nodo padre
            dati_beni = self.xml_doc.getroot().xpath("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi", namespaces=self.NS)[0]
            
            # Crea una nuova linea
            new_line = etree.Element("DettaglioLinee")
            etree.SubElement(new_line, "NumeroLinea").text = "0"  # Sarà aggiornato dopo
            etree.SubElement(new_line, "Descrizione").text = "Nuovo articolo"
            etree.SubElement(new_line, "Quantita").text = "1.0000000"
            etree.SubElement(new_line, "UnitaMisura").text = "NR"
            etree.SubElement(new_line, "PrezzoUnitario").text = "0.0000000"
            etree.SubElement(new_line, "PrezzoTotale").text = "0.0000000"
            etree.SubElement(new_line, "AliquotaIVA").text = "22.00"
            
            # Aggiorna i dati delle linee per avere informazioni aggiornate
            self.refresh_lines_data()
            
            # Decide dove inserire la linea:
            # 1. Se c'è una linea CONAI, inserisci prima della linea CONAI
            # 2. Se non c'è una linea CONAI, inserisci prima di DatiRiepilogo
            # 3. Altrimenti, inserisci alla fine di DatiBeniServizi
            
            if self.conai_line is not None:
                dati_beni.insert(dati_beni.index(self.conai_line), new_line)
                self.log("Inserita nuova linea prima della linea CONAI")
            else:
                riepilogo = dati_beni.xpath("./DatiRiepilogo", namespaces=self.NS)
                if riepilogo:
                    dati_beni.insert(dati_beni.index(riepilogo[0]), new_line)
                    self.log("Inserita nuova linea prima di DatiRiepilogo")
                else:
                    dati_beni.append(new_line)
                    self.log("Inserita nuova linea alla fine di DatiBeniServizi")
            
            # Aggiorna i numeri delle linee
            self.update_line_numbers()
            
            # Aggiorna i dati e l'interfaccia
            self.refresh_lines_data()
            
            # Posizionati sulla nuova linea
            # Dato che la linea è stata aggiunta alla fine delle linee normali
            self.current_line_index = self.total_lines - 1
            
            self.update_line_fields()
            self.update_nav_buttons()
            
            self.log(f"Aggiunta nuova linea in posizione {self.current_line_index + 1}")
        except Exception as e:
            self.log(f"Errore nell'aggiunta di una nuova linea: {str(e)}")
            traceback.print_exc()
            messagebox.showerror("Errore", f"Impossibile aggiungere una nuova linea:\n{str(e)}")
        self.update_riepilogo_totals()

            
    def update_line_numbers(self):
        try:
            # Ottieni tutte le linee di dettaglio
            linee = self.xml_doc.getroot().xpath("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi/DettaglioLinee", namespaces=self.NS)
            
            # Aggiorna il numero di ogni linea
            for i, linea in enumerate(linee, 1):  # inizia da 1
                numero_linea = linea.find("NumeroLinea")
                if numero_linea is not None:
                    numero_linea.text = str(i)
                else:
                    etree.SubElement(linea, "NumeroLinea").text = str(i)
            
            self.log(f"Numeri delle linee aggiornati (1-{len(linee)})")
        except Exception as e:
            self.log(f"Errore nell'aggiornamento dei numeri delle linee: {str(e)}")
            traceback.print_exc()


    def refresh_lines_data(self):
        """
        Aggiorna i dati delle linee dal documento XML, escludendo la linea CONAI
        """
        try:
            root = self.xml_doc.getroot()
            all_lines = root.xpath("//p:FatturaElettronica/FatturaElettronicaBody/DatiBeniServizi/DettaglioLinee", namespaces=self.NS)
            
            # Dividi tra linee normali e linea CONAI
            self.normal_lines = []
            self.conai_line = None
            
            for line in all_lines:
                descr = line.find("Descrizione")
                if descr is not None and descr.text and "CONTRIBUTO CONAI ASSOLTO" in descr.text.upper():
                    self.conai_line = line
                else:
                    self.normal_lines.append(line)
            
            self.total_lines = len(self.normal_lines)
            
            # Aggiorna lo stato della checkbox CONAI solo se esiste
            if hasattr(self, 'conai_var'):
                self.conai_var.set(self.conai_line is not None)
            
            # Assicurati che l'indice corrente sia valido
            if self.current_line_index >= self.total_lines:
                self.current_line_index = max(0, self.total_lines - 1)
            self.log(f"Aggiornati dati: {self.total_lines} linee normali, CONAI: {'presente' if self.conai_line is not None else 'assente'}")    
        except Exception as e:
            self.log(f"Errore nell'aggiornamento dei dati delle linee: {str(e)}")
            traceback.print_exc()

    def toggle_conai(self):
        """Gestisce l'attivazione/disattivazione della linea CONAI tramite checkbox"""
        if self.conai_var.get():
            self.log("Attivazione linea CONTRIBUTO CONAI ASSOLTO")
            self.add_conai_line()
        else:
            self.log("Rimozione linea CONTRIBUTO CONAI ASSOLTO")
            self.remove_conai_line()

    def update_riepilogo_totals(self):
        """
        Aggiorna automaticamente i totali nei dati di riepilogo basati sulle linee di dettaglio.
        """
        try:
            # Calcola la somma dei prezzi totali di tutte le linee
            total_amount = 0.0
            
            # Aggiorna i dati per assicurarci di avere le informazioni più recenti
            self.refresh_lines_data()
            
            # Somma i prezzi totali di tutte le linee normali
            for line in self.normal_lines:
                prezzo_totale_elem = line.find("PrezzoTotale")
                if prezzo_totale_elem is not None and prezzo_totale_elem.text:
                    try:
                        total_amount += float(prezzo_totale_elem.text)
                        self.log(f"Aggiunto prezzo: {prezzo_totale_elem.text}, totale parziale: {total_amount}")
                    except ValueError:
                        self.log(f"Errore nel convertire il prezzo totale: {prezzo_totale_elem.text}")
            
            # Formatta l'importo totale con 2 decimali
            total_amount_formatted = f"{total_amount:.2f}"
            self.log(f"Imponibile calcolato: {total_amount_formatted}")
            
            # Aggiorna direttamente nei campi dell'interfaccia se esistono
            for xpath, field_data in self.edit_fields.items():
                if "ImponibileImporto" in xpath and field_data["widget"].winfo_exists():
                    field_data["widget"].delete(0, tk.END)
                    field_data["widget"].insert(0, total_amount_formatted)
                    self.log(f"Aggiornato campo UI imponibile: {total_amount_formatted}")
                    
                    # Aggiorna anche l'elemento XML
                    if field_data["element"] is not None:
                        field_data["element"].text = total_amount_formatted
                        self.log(f"Aggiornato elemento XML imponibile: {total_amount_formatted}")
            
            # Calcola anche l'imposta e il totale documento
            aliquota_iva = 22.0  # Valore predefinito se non troviamo l'elemento
            
            # Cerca l'aliquota IVA nel documento
            for xpath, field_data in self.edit_fields.items():
                if "AliquotaIVA" in xpath and "DatiRiepilogo" in xpath and field_data["widget"].winfo_exists():
                    try:
                        aliquota_iva = float(field_data["widget"].get())
                        self.log(f"Trovata aliquota IVA: {aliquota_iva}%")
                        break
                    except ValueError:
                        pass
            
            # Calcola l'imposta
            imposta = total_amount * (aliquota_iva / 100.0)
            imposta_formatted = f"{imposta:.2f}"
            
            # Aggiorna il campo imposta nell'interfaccia
            for xpath, field_data in self.edit_fields.items():
                if "Imposta" in xpath and field_data["widget"].winfo_exists():
                    field_data["widget"].delete(0, tk.END)
                    field_data["widget"].insert(0, imposta_formatted)
                    self.log(f"Aggiornato campo UI imposta: {imposta_formatted}")
                    
                    # Aggiorna anche l'elemento XML
                    if field_data["element"] is not None:
                        field_data["element"].text = imposta_formatted
                        self.log(f"Aggiornato elemento XML imposta: {imposta_formatted}")
            
            # Calcola e aggiorna l'importo totale documento
            importo_totale = total_amount + imposta
            importo_totale_formatted = f"{importo_totale:.2f}"
            
            for xpath, field_data in self.edit_fields.items():
                if "ImportoTotaleDocumento" in xpath and field_data["widget"].winfo_exists():
                    field_data["widget"].delete(0, tk.END)
                    field_data["widget"].insert(0, importo_totale_formatted)
                    self.log(f"Aggiornato campo UI importo totale: {importo_totale_formatted}")
                    
                    # Aggiorna anche l'elemento XML
                    if field_data["element"] is not None:
                        field_data["element"].text = importo_totale_formatted
                        self.log(f"Aggiornato elemento XML importo totale: {importo_totale_formatted}")
            
            # Aggiorna anche il campo ImportoPagamento nei dati pagamento
            for xpath, field_data in self.edit_fields.items():
                if "ImportoPagamento" in xpath and field_data["widget"].winfo_exists():
                    field_data["widget"].delete(0, tk.END)
                    field_data["widget"].insert(0, importo_totale_formatted)
                    self.log(f"Aggiornato campo UI importo pagamento: {importo_totale_formatted}")
                    
                    # Aggiorna anche l'elemento XML
                    if field_data["element"] is not None:
                        field_data["element"].text = importo_totale_formatted
                        self.log(f"Aggiornato elemento XML importo pagamento: {importo_totale_formatted}")
            
        except Exception as e:
            self.log(f"Errore nell'aggiornamento dei totali: {str(e)}")
            traceback.print_exc() 



    def calcola_imposta(self, event=None):
        """
        Metodo per calcolare l'imposta in base all'aliquota e all'imponibile.
        Può essere chiamato sia manualmente che dagli eventi dell'interfaccia.
        """
        try:
            # Verifico che i widget siano disponibili
            if not all([self.aliquota_iva_widget, self.imponibile_widget, self.imposta_widget]):
                self.log("Widget non disponibili per il calcolo dell'imposta")
                return
                
            # Ottieni i valori dai campi
            aliquota_str = self.aliquota_iva_widget.get()
            imponibile_str = self.imponibile_widget.get().replace(',', '.')
            
            # Log per debug
            self.log(f"Calcolando imposta con aliquota: {aliquota_str}, imponibile: {imponibile_str}")
            
            # Verifica che i valori non siano vuoti
            if aliquota_str and imponibile_str:
                aliquota = float(aliquota_str)
                imponibile = float(imponibile_str)
                
                # Calcola l'imposta
                imposta = (imponibile * aliquota) / 100.0
                
                # Formatta il risultato con 2 decimali
                imposta_formatted = f"{imposta:.2f}"
                
                # Aggiorna il campo imposta
                self.imposta_widget.delete(0, tk.END)
                self.imposta_widget.insert(0, imposta_formatted)
                
                # Aggiorna anche l'elemento XML corrispondente
                for xpath, field_data in self.edit_fields.items():
                    if "Imposta" in xpath and field_data["element"] is not None:
                        field_data["element"].text = imposta_formatted
                        break
                
                # Aggiorna anche l'importo totale documento
                self.update_riepilogo_totals()
                
                self.log(f"Imposta ricalcolata: {imposta_formatted}")
        except Exception as e:
            self.log(f"Errore nel calcolo dell'imposta: {str(e)}")
            traceback.print_exc()   

    def load_template(self):
        """Carica un modello di fattura precompilato dalla cartella del progetto"""
        try:
            template_path = os.path.join(self.project_dir, "modelloFattura.xml")
            
            # Verifica che il file esista
            if not os.path.exists(template_path):
                self.log(f"Errore: File modello non trovato in {template_path}")
                messagebox.showerror("Errore", f"File modello non trovato.\nPercorso cercato: {template_path}")
                return
                
            # Carica il file modello
            self.xml_path = template_path
            self.xml_label.config(text=os.path.basename(template_path))
            self.log(f"File modello caricato: {template_path}")
            
            try:
                self.xml_doc = etree.parse(template_path)
                self.log("File modello XML caricato con successo")
                messagebox.showinfo("Modello Caricato", "Il modello di fattura è stato caricato con successo.")
                # Aggiorna lo stato dei pulsanti
                self.update_button_states()
            except Exception as e:
                self.log(f"Errore nel caricamento del file modello XML: {str(e)}")
                self.xml_doc = None
                # Aggiorna lo stato dei pulsanti (disabilita)
                self.update_button_states()
                messagebox.showerror("Errore", f"Errore nel caricamento del file modello XML:\n{str(e)}")
        except Exception as e:
            self.log(f"Errore nel caricamento del modello: {str(e)}")
            messagebox.showerror("Errore", f"Errore nel caricamento del modello:\n{str(e)}")    
            
    def export_to_excel(self):
        """Esporta i dati del file XML attuale in Excel"""
        if not self.xml_path or not self.xml_doc:
            messagebox.showerror("Errore", "Seleziona prima un file XML valido")
            return
        
        # Verifica se è stato impostato un file Excel
        if not self.excel_manager.excel_path:
            # Chiedi all'utente se vuole creare un nuovo file o caricarne uno esistente
            choice = messagebox.askquestion("Database Excel",
                                        "Non è stato specificato un database Excel.\n\n" +
                                        "Vuoi crearne uno nuovo?\n\n" +
                                        "Seleziona 'Sì' per creare un nuovo database Excel, " +
                                        "'No' per selezionare un database esistente.")
            
            if choice == 'yes':
                # Crea un nuovo database Excel
                self.create_excel_db()
            else:
                # Carica un database esistente
                self.load_excel_db()
        
        # Controlla nuovamente se dopo le azioni dell'utente c'è un percorso Excel valido
        if not self.excel_manager.excel_path:
            self.log("Operazione annullata: nessun database Excel specificato")
            return
        
        try:
            success = self.excel_manager.export_xml_to_excel(self.xml_doc)
            if success:
                messagebox.showinfo("Esportazione completata", 
                                f"I dati sono stati esportati con successo nel file Excel")
            else:
                messagebox.showerror("Errore", 
                                "Si è verificato un errore durante l'esportazione in Excel")
        except Exception as e:
            self.log(f"Errore nell'esportazione in Excel: {str(e)}")
            traceback.print_exc()
            messagebox.showerror("Errore", f"Errore nell'esportazione in Excel:\n{str(e)}")


    def create_xml_from_excel(self):
        """Crea un nuovo file XML dai dati in Excel"""
        try:
            success, new_xml_path = self.excel_manager.import_excel_to_xml()
            
            if success and new_xml_path:
                # Chiedi all'utente se vuole caricare il nuovo file XML
                result = messagebox.askyesno("File XML creato", 
                                            f"Il file XML è stato creato con successo.\n\nVuoi caricarlo ora?")
                if result:
                    # Carica il nuovo file XML
                    self.xml_path = new_xml_path
                    self.xml_label.config(text=os.path.basename(new_xml_path))
                    self.log(f"File XML creato e caricato: {new_xml_path}")
                    
                    try:
                        self.xml_doc = etree.parse(new_xml_path)
                        self.log("Nuovo file XML caricato con successo")
                        # Aggiorna lo stato dei pulsanti
                        self.update_button_states()
                    except Exception as e:
                        self.log(f"Errore nel caricamento del nuovo file XML: {str(e)}")
                        self.xml_doc = None
                        # Aggiorna lo stato dei pulsanti (disabilita)
                        self.update_button_states()
        except Exception as e:
            self.log(f"Errore nella creazione del file XML da Excel: {str(e)}")
            traceback.print_exc()
            messagebox.showerror("Errore", f"Errore nella creazione del file XML da Excel:\n{str(e)}")


    def manage_invoices(self):
        """Mostra una finestra per gestire le fatture in Excel"""
        try:
            # Verifica se è stato impostato un file Excel
            if not self.excel_manager.excel_path:
                # Chiedi all'utente se vuole creare un nuovo file o caricarne uno esistente
                choice = messagebox.askquestion("Database Excel",
                                            "Non è stato specificato un database Excel.\n\n" +
                                            "Vuoi crearne uno nuovo?\n\n" +
                                            "Seleziona 'Sì' per creare un nuovo database Excel, " +
                                            "'No' per selezionare un database esistente.")
                
                if choice == 'yes':
                    # Crea un nuovo database Excel
                    self.create_excel_db()
                else:
                    # Carica un database esistente
                    self.load_excel_db()
            
            # Controlla nuovamente se dopo le azioni dell'utente c'è un percorso Excel valido
            if not self.excel_manager.excel_path:
                self.log("Operazione annullata: nessun database Excel specificato")
                return
            
            # Verifica che il file Excel esista
            excel_path = self.excel_manager.excel_path
            if not os.path.exists(excel_path):
                messagebox.showinfo("Informazione", 
                                f"Il file Excel non esiste.\nCrea prima una fattura in Excel.")
                return
            
            # Ottieni l'elenco delle fatture
            invoices = self.excel_manager.list_invoices()
            if not invoices:
                messagebox.showinfo("Informazione", "Nessuna fattura trovata nel file Excel")
                return
            
            # Crea una finestra di dialogo per la gestione
            manager = tk.Toplevel(self)
            manager.title("Gestione Fatture in Excel")
            manager.geometry("800x500")
            manager.transient(self)
            manager.grab_set()
            
            # Frame principale
            main_frame = tk.Frame(manager)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Intestazione
            header = tk.Label(main_frame, text="Fatture salvate in Excel", font=("", 12, "bold"))
            header.pack(fill=tk.X, pady=(0, 10))
            
            # Frame per la tabella con scrollbar
            table_frame = tk.Frame(main_frame)
            table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            
            # Scrollbar
            scrollbar = tk.Scrollbar(table_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Treeview per visualizzare le fatture
            columns = ("id", "numero", "data", "cedente", "cessionario")
            tree = ttk.Treeview(table_frame, columns=columns, show='headings', yscrollcommand=scrollbar.set)
            
            # Configurazione colonne
            tree.heading("id", text="ID")
            tree.heading("numero", text="Numero Fattura")
            tree.heading("data", text="Data")
            tree.heading("cedente", text="Cedente")
            tree.heading("cessionario", text="Cessionario")
            
            tree.column("id", width=80, anchor="w")
            tree.column("numero", width=100, anchor="w")
            tree.column("data", width=100, anchor="w")
            tree.column("cedente", width=200, anchor="w")
            tree.column("cessionario", width=200, anchor="w")
            
            # Nascondi la colonna ID
            tree.column("id", width=0, stretch=tk.NO)
            
            # Popola la tabella
            for invoice in invoices:
                tree.insert("", tk.END, values=invoice)
            
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=tree.yview)
            
            # Frame per i pulsanti
            button_frame = tk.Frame(main_frame)
            button_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Funzioni per le azioni sui pulsanti
            def create_xml():
                selection = tree.selection()
                if not selection:
                    messagebox.showwarning("Attenzione", "Seleziona prima una fattura")
                    return
                
                item = tree.item(selection[0])
                invoice_id = item["values"][0]
                
                # Chiedi il percorso di salvataggio
                output_path = filedialog.asksaveasfilename(
                    title="Salva il nuovo file XML",
                    defaultextension=".xml",
                    filetypes=[("File XML", "*.xml")]
                )
                if not output_path:
                    return
                
                # Crea il file XML
                success, new_xml_path = self.excel_manager.create_xml_from_excel_by_id(invoice_id, output_path)
                
                if success:
                    result = messagebox.askyesno("File XML creato", 
                                            f"Il file XML è stato creato con successo.\n\nVuoi caricarlo ora?")
                    if result:
                        # Carica il nuovo file XML
                        self.xml_path = new_xml_path
                        self.xml_label.config(text=os.path.basename(new_xml_path))
                        self.log(f"File XML creato e caricato: {new_xml_path}")
                        
                        try:
                            self.xml_doc = etree.parse(new_xml_path)
                            self.log("Nuovo file XML caricato con successo")
                            # Aggiorna lo stato dei pulsanti
                            self.update_button_states()
                            manager.destroy()  # Chiudi la finestra
                        except Exception as e:
                            self.log(f"Errore nel caricamento del nuovo file XML: {str(e)}")
                            self.xml_doc = None
                            # Aggiorna lo stato dei pulsanti
                            self.update_button_states()

            def edit_invoice_from_excel():
                """Modifica la fattura selezionata nel treeview"""
                selection = tree.selection()
                if not selection:
                    messagebox.showwarning("Attenzione", "Seleziona prima una fattura")
                    return
                
                item = tree.item(selection[0])
                invoice_id = item["values"][0]
                invoice_number = item["values"][1]
                
                # Genera un file XML temporaneo per la modifica
                import tempfile
                temp_dir = tempfile.gettempdir()
                temp_file = os.path.join(temp_dir, f"temp_invoice_{invoice_number}.xml")
                
                # Controlla se il metodo create_xml_from_excel_by_id esiste
                if hasattr(self.excel_manager, 'create_xml_from_excel_by_id'):
                    # Usa il metodo se esiste
                    success, xml_path = self.excel_manager.create_xml_from_excel_by_id(invoice_id, temp_file)
                else:
                    # Altrimenti usa il metodo import_excel_to_xml con id e path specificati manualmente
                    self.log("Utilizzo metodo alternativo per generare XML")
                    try:
                        # Carica il workbook Excel
                        import openpyxl
                        excel_path = self.excel_manager.excel_path
                        wb = openpyxl.load_workbook(excel_path)
                        
                        # Estrai i dati della fattura
                        invoice_data = self.excel_manager._get_invoice_data_by_id(wb, invoice_id)
                        if not invoice_data:
                            messagebox.showerror("Errore", f"Dati non trovati per la fattura con ID: {invoice_id}")
                            return
                        
                        # Genera XML dalla fattura
                        xml_doc = self.excel_manager._generate_xml_from_invoice_data(invoice_data)
                        
                        # Salva il file XML
                        xml_string = etree.tostring(xml_doc, pretty_print=True, encoding="UTF-8", 
                                                    xml_declaration=True).decode("utf-8")
                        
                        # Aggiungi gli stessi miglioramenti del metodo create_xml_from_excel_by_id
                        xml_string = xml_string.replace("<FatturaElettronica ", "<p:FatturaElettronica ")
                        xml_string = xml_string.replace("</FatturaElettronica>", "</p:FatturaElettronica>")
                        xml_string = xml_string.replace("<p:FatturaElettronica ", 
                                                    "<p:FatturaElettronica xmlns:p=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\" ")
                        xml_string = xml_string.replace(" xmlns=\"http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2\"", "")
                        
                        # Aggiungi riferimento stylesheet
                        stylesheet_ref = '<?xml-stylesheet type="text/xsl" href="./fatturapa_v1.2_asw.xsl"?>\n'
                        if '<?xml ' in xml_string:
                            xml_decl_end = xml_string.find('?>') + 2
                            xml_string = xml_string[:xml_decl_end] + '\n' + stylesheet_ref + xml_string[xml_decl_end:]
                        else:
                            xml_string = stylesheet_ref + xml_string
                        
                        with open(temp_file, 'w', encoding='utf-8') as f:
                            f.write(xml_string)
                        
                        success = True
                        xml_path = temp_file
                        
                    except Exception as e:
                        self.log(f"Errore nella generazione del XML: {str(e)}")
                        traceback.print_exc()
                        messagebox.showerror("Errore", f"Errore nella generazione del XML:\n{str(e)}")
                        success = False
                        xml_path = None
                
                if not success:
                    messagebox.showerror("Errore", "Impossibile creare il file XML temporaneo per la modifica")
                    return
                
                # Carica il file XML appena creato
                self.xml_path = xml_path
                self.xml_label.config(text=f"Temp: {os.path.basename(xml_path)}")
                self.log(f"File XML temporaneo creato per modifica: {xml_path}")
                
                try:
                    self.xml_doc = etree.parse(xml_path)
                    self.log("File XML temporaneo caricato con successo")
                    
                    # Chiudi la finestra di gestione
                    manager.destroy()
                    
                    # Aggiorna lo stato dei pulsanti
                    self.update_button_states()
                    
                    # Avvia la modalità di modifica
                    self.edit_invoice()
                    
                except Exception as e:
                    self.log(f"Errore nel caricamento del file XML temporaneo: {str(e)}")
                    self.xml_doc = None
                    self.update_button_states()
                    messagebox.showerror("Errore", f"Errore nel caricamento del file XML:\n{str(e)}")
            
            def delete_invoice():
                selection = tree.selection()
                if not selection:
                    messagebox.showwarning("Attenzione", "Seleziona prima una fattura")
                    return
                
                item = tree.item(selection[0])
                invoice_id = item["values"][0]
                invoice_numero = item["values"][1]
                
                # Chiedi conferma
                result = messagebox.askyesno("Conferma eliminazione", 
                                            f"Sei sicuro di voler eliminare la fattura {invoice_numero}?")
                if not result:
                    return
                
                # Elimina la fattura
                success = self.excel_manager.delete_invoice(invoice_id)
                
                if success:
                    messagebox.showinfo("Eliminazione completata", 
                                    f"La fattura {invoice_numero} è stata eliminata con successo")
                    # Rimuovi dalla tabella
                    tree.delete(selection[0])
                else:
                    messagebox.showerror("Errore", 
                                    f"Si è verificato un errore durante l'eliminazione della fattura")
            
            def open_excel():
                # Apri il file Excel con l'applicazione predefinita
                try:
                    excel_path = self.excel_manager.excel_path
                    if os.path.exists(excel_path):
                        os.startfile(excel_path)
                    else:
                        messagebox.showinfo("Informazione", "File Excel non trovato")
                except Exception as e:
                    messagebox.showerror("Errore", f"Impossibile aprire il file Excel:\n{str(e)}")
            
            def refresh_list():
                # Aggiorna l'elenco delle fatture
                tree.delete(*tree.get_children())
                invoices = self.excel_manager.list_invoices()
                for invoice in invoices:
                    tree.insert("", tk.END, values=invoice)
            
            # Pulsanti azioni
            create_btn = tk.Button(button_frame, text="Crea XML", command=create_xml,
                                bg="#4CAF50", fg="white", width=15, padx=5, pady=5)
            create_btn.pack(side=tk.LEFT, padx=(0, 5))

            # Pulsante Modifica
            edit_btn = tk.Button(button_frame, text="Modifica", command=edit_invoice_from_excel,
                            bg="#2196F3", fg="white", width=15, padx=5, pady=5)
            edit_btn.pack(side=tk.LEFT, padx=5)
            
            delete_btn = tk.Button(button_frame, text="Elimina", command=delete_invoice,
                                bg="#F44336", fg="white", width=15, padx=5, pady=5)
            delete_btn.pack(side=tk.LEFT, padx=5)
            
            excel_btn = tk.Button(button_frame, text="Apri Excel", command=open_excel,
                                bg="#673AB7", fg="white", width=15, padx=5, pady=5)
            excel_btn.pack(side=tk.LEFT, padx=5)
            
            refresh_btn = tk.Button(button_frame, text="Aggiorna", command=refresh_list,
                                bg="#FF9800", fg="white", width=15, padx=5, pady=5)
            refresh_btn.pack(side=tk.LEFT, padx=5)
            
            close_btn = tk.Button(button_frame, text="Chiudi", command=manager.destroy,
                                width=15, padx=5, pady=5)
            close_btn.pack(side=tk.RIGHT, padx=5)
            
            # Gestione doppio click su riga
            tree.bind("<Double-1>", lambda e: edit_invoice_from_excel())  # Doppio click modifica la fattura
            
        except Exception as e:
            self.log(f"Errore nella gestione delle fatture in Excel: {str(e)}")
            traceback.print_exc()
            messagebox.showerror("Errore", f"Errore nella gestione delle fatture in Excel:\n{str(e)}")

            
    # Aggiorna il metodo di aggiornamento dello stato dei pulsanti
    def update_button_states(self):
        """Aggiorna lo stato dei pulsanti in base alle condizioni attuali"""
        
        # Controlla se un file XML è caricato
        xml_loaded = self.xml_path is not None and self.xml_doc is not None
        
        # Controlla se un foglio di stile XSL è selezionato
        xsl_selected = self.xsl_path is not None
        
        # Imposta lo stato dei pulsanti in base alle condizioni
        if xml_loaded:
            self.edit_btn.config(state=tk.NORMAL)
            self.excel_save_btn.config(state=tk.NORMAL)
        else:
            self.edit_btn.config(state=tk.DISABLED)
            self.excel_save_btn.config(state=tk.DISABLED)
        
        # Il pulsante Visualizza richiede sia XML che XSL
        if xml_loaded and xsl_selected:
            self.view_btn.config(state=tk.NORMAL)
        else:
            self.view_btn.config(state=tk.DISABLED)
        
        # Aggiorna anche le informazioni del database Excel
        self.update_excel_db_info()


    # Aggiungi questi nuovi metodi alla classe FatturaViewer


    # Modifica i metodi load_excel_db e create_excel_db per aggiornare l'interfaccia
    def load_excel_db(self):
        """Carica un database Excel esistente"""
        try:
            # Chiedi all'utente di selezionare il file Excel
            filepath = filedialog.askopenfilename(
                title="Seleziona il file Excel di database",
                filetypes=[("File Excel", "*.xlsx")]
            )
            
            if not filepath:
                return  # L'utente ha annullato la selezione
            
            # Verifica che il file esista e sia un file Excel valido
            if not os.path.exists(filepath):
                messagebox.showerror("Errore", "Il file selezionato non esiste.")
                return
                
            # Aggiorna il percorso del file Excel nel manager
            self.excel_manager.excel_path = filepath
            
            # Verifica che il file Excel abbia i fogli necessari
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                
                # Controlla se i fogli richiesti esistono
                required_sheets = [
                    self.excel_manager.master_sheet_name,
                    self.excel_manager.details_sheet_name,
                    self.excel_manager.summary_sheet_name,
                    self.excel_manager.structure_sheet_name
                ]
                
                missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
                
                if missing_sheets:
                    # Se mancano fogli, chiedi all'utente se vuole crearli
                    result = messagebox.askyesno("Fogli mancanti", 
                        f"Il file Excel non contiene tutti i fogli necessari. Mancano: {', '.join(missing_sheets)}.\n\n" +
                        "Vuoi creare i fogli mancanti?")
                    
                    if result:
                        # Crea i fogli mancanti
                        wb.close()  # Chiudi il file read-only
                        self.create_excel_sheets(filepath, missing_sheets)
                        self.log(f"Fogli mancanti creati nel file: {filepath}")
                    else:
                        self.log("Operazione annullata: fogli mancanti non creati")
                        return
                else:
                    wb.close()  # Chiudi il file read-only
                    
                self.log(f"Database Excel caricato: {filepath}")
                messagebox.showinfo("Database Excel", f"Database Excel caricato con successo:\n{os.path.basename(filepath)}")
                
                # Aggiorna l'interfaccia utente
                self.update_button_states()
                
            except Exception as e:
                self.log(f"Errore nel controllo del file Excel: {str(e)}")
                messagebox.showerror("Errore", f"Errore nel controllo del file Excel:\n{str(e)}")
                return
            
        except Exception as e:
            self.log(f"Errore nel caricamento del database Excel: {str(e)}")
            messagebox.showerror("Errore", f"Errore nel caricamento del database Excel:\n{str(e)}")

    def create_excel_db(self):
        """Crea un nuovo database Excel vuoto"""
        try:
            # Chiedi all'utente dove salvare il nuovo file Excel
            filepath = filedialog.asksaveasfilename(
                title="Salva il nuovo database Excel",
                filetypes=[("File Excel", "*.xlsx")],
                defaultextension=".xlsx"
            )
            
            if not filepath:
                return  # L'utente ha annullato la selezione
            
            # Assicurati che il file abbia estensione .xlsx
            if not filepath.lower().endswith('.xlsx'):
                filepath += '.xlsx'
            
            # Crea un nuovo file Excel con i fogli necessari
            self.create_excel_sheets(filepath, [
                self.excel_manager.master_sheet_name,
                self.excel_manager.details_sheet_name, 
                self.excel_manager.summary_sheet_name,
                self.excel_manager.structure_sheet_name
            ])
            
            # Aggiorna il percorso del file Excel nel manager
            self.excel_manager.excel_path = filepath
            
            self.log(f"Nuovo database Excel creato: {filepath}")
            messagebox.showinfo("Database Excel", f"Nuovo database Excel creato con successo:\n{os.path.basename(filepath)}")
            
            # Aggiorna l'interfaccia utente
            self.update_button_states()
            
        except Exception as e:
            self.log(f"Errore nella creazione del database Excel: {str(e)}")
            messagebox.showerror("Errore", f"Errore nella creazione del database Excel:\n{str(e)}")



    def create_excel_sheets(self, filepath, sheet_names):
        """
        Crea i fogli necessari in un file Excel esistente o nuovo
        
        Args:
            filepath: Percorso del file Excel
            sheet_names: Lista dei nomi dei fogli da creare
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Verifica se il file esiste già
        file_exists = os.path.exists(filepath)
        
        if file_exists:
            # Apri il file esistente
            wb = openpyxl.load_workbook(filepath)
        else:
            # Crea un nuovo workbook
            wb = openpyxl.Workbook()
            # Rimuovi il foglio di default
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        
        # Crea i fogli mancanti con le intestazioni appropriate
        for sheet_name in sheet_names:
            # Salta se il foglio esiste già
            if sheet_name in wb.sheetnames:
                continue
                
            # Crea un nuovo foglio
            sheet = wb.create_sheet(title=sheet_name)
            
            # Stile per le intestazioni
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Aggiungi le intestazioni in base al tipo di foglio
            if sheet_name == self.excel_manager.master_sheet_name:
                headers = [
                    "ID_Fattura", "NumeroFattura", "DataFattura", "TipoDocumento", 
                    "ImportoTotale", "CedenteDenominazione", "CedentePartitaIVA", 
                    "CessionarioDenominazione", "CessionarioPartitaIVA", "NotaFattura"
                ]
            elif sheet_name == self.excel_manager.details_sheet_name:
                headers = [
                    "ID_Fattura", "NumeroLinea", "Descrizione", "Quantita", 
                    "UnitaMisura", "PrezzoUnitario", "PrezzoTotale", "AliquotaIVA", "Note"
                ]
            elif sheet_name == self.excel_manager.summary_sheet_name:
                headers = [
                    "ID_Fattura", "AliquotaIVA", "ImponibileImporto", "Imposta", 
                    "EsigibilitaIVA", "Natura"
                ]
            elif sheet_name == self.excel_manager.structure_sheet_name:
                headers = [
                    "TagXML", "Percorso", "Descrizione"
                ]
            else:
                # Foglio generico senza intestazioni specifiche
                headers = []
            
            # Aggiungi le intestazioni al foglio
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Imposta larghezza colonne
            for i, _ in enumerate(headers, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
        
        # Salva il workbook
        wb.save(filepath)



    # Aggiungi un metodo per aggiornare le informazioni del database Excel
    def update_excel_db_info(self):
        """Aggiorna l'etichetta del database Excel e lo stato dei pulsanti correlati"""
        if self.excel_manager.excel_path and os.path.exists(self.excel_manager.excel_path):
            # Mostra il nome del file (non il percorso completo per evitare etichette troppo lunghe)
            db_name = os.path.basename(self.excel_manager.excel_path)
            self.excel_db_label.config(text=db_name, fg="black")
            
            # Abilita il pulsante Gestisci Fatture
            self.excel_manage_btn.config(state=tk.NORMAL)
        else:
            # Nessun database specificato o il file non esiste
            self.excel_db_label.config(text="Non specificato", fg="gray")
            
            # Disabilita il pulsante Gestisci Fatture
            self.excel_manage_btn.config(state=tk.DISABLED)
                    
if __name__ == "__main__":
    app = FatturaViewer()
    app.mainloop()
